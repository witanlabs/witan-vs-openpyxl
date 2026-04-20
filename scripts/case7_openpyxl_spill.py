"""Task: open report_spillref.xlsx with openpyxl, report the formulas in
Summary!F2, G2, H2 (which consume Summary!D2# via A1# spill syntax), and
attempt to tokenize them."""
import sys
from openpyxl import load_workbook
from openpyxl.formula import Tokenizer

SRC = sys.argv[1]

print("=== load_workbook ===")
try:
    wb = load_workbook(SRC)
    print("  loaded OK")
except Exception as e:
    print(f"  FAIL ({type(e).__name__}): {e}")
    sys.exit(1)

ws = wb["Summary"]
for addr in ("D2", "F2", "G2", "H2"):
    c = ws[addr]
    print(f"\n{addr}: value={c.value!r}")
    if isinstance(c.value, str) and c.value.startswith("="):
        print(f"  Tokenizer({c.value!r})")
        try:
            t = Tokenizer(c.value)
            for tok in t.items:
                print(f"    type={tok.type!r} subtype={tok.subtype!r} value={tok.value!r}")
        except Exception as e:
            print(f"    Tokenizer FAIL ({type(e).__name__}): {e}")

print()
print("=== Round-trip: save and reload ===")
OUT = "/tmp/case7_spill_roundtrip.xlsx"
wb.save(OUT)
wb2 = load_workbook(OUT)
for addr in ("D2", "F2", "G2", "H2"):
    print(f"  after save: {addr} = {wb2['Summary'][addr].value!r}")
