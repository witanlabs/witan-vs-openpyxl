"""Case 18: recalculate a two-variable What-If Data Table.

The fixture is the existing Case 8 data-table workbook. The task changes a
normal model input, Model!B4 (fixed cost), from 5000 to 6000 and then reports
the updated data-table values. Witan is the control because setCells updates the
formula cell and the data table. openpyxl writes the input but cannot calculate.
LibreOffice Calc opens and rewrites the data-table body to TABLE(...) formulas,
but does not save calculated table-body values.
"""
from __future__ import annotations

import json
import os
import shlex
import shutil
import subprocess
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
FIXTURE = ROOT / "fixtures" / "sensitivity2d.xlsx"
OUT = ROOT / "outputs" / "whatif_datatable_recalc"
SOFFICE = Path(os.environ.get("SOFFICE", "/opt/homebrew/bin/soffice"))
WITAN_CMD = shlex.split(os.environ.get("WITAN_CMD", "npx witan"))
KEY_CELLS = ["D1", "E2", "F3", "G4", "H5", "I6"]
EXPECTED = {"D1": 14000, "E2": 12800, "F3": 30000, "G4": 52200, "H5": 79400, "I6": 111600}


def run(cmd: list[str]) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        cmd,
        cwd=ROOT,
        check=True,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
    )


def read_cells(path: Path, data_only: bool) -> dict[str, object]:
    wb = load_workbook(path, data_only=data_only)
    ws = wb["Model"]
    return {addr: ws[addr].value for addr in KEY_CELLS}


def resave_with_libreoffice(src: Path) -> Path:
    if not SOFFICE.exists():
        raise RuntimeError(f"LibreOffice not found at {SOFFICE}; set SOFFICE=/path/to/soffice")

    work = OUT / "_lo_work"
    converted = OUT / "_lo_out"
    profile = OUT / "_lo_profile"
    shutil.rmtree(work, ignore_errors=True)
    shutil.rmtree(converted, ignore_errors=True)
    shutil.rmtree(profile, ignore_errors=True)
    work.mkdir(parents=True)
    converted.mkdir(parents=True)
    profile.mkdir(parents=True)

    local = work / src.name
    shutil.copy(src, local)
    run(
        [
            str(SOFFICE),
            "--headless",
            f"-env:UserInstallation={profile.resolve().as_uri()}",
            "--convert-to",
            "xlsx",
            "--outdir",
            str(converted),
            str(local),
        ]
    )
    out = converted / local.name
    if not out.exists():
        raise RuntimeError(f"LibreOffice did not produce {out}")
    return out


def run_witan() -> dict[str, object]:
    dst = OUT / "case18_witan.xlsx"
    shutil.copy(FIXTURE, dst)
    code = (
        'const r = await xlsx.setCells(wb, [{address:"Model!B4", value:6000}]); '
        f"return {json.dumps(KEY_CELLS)}.reduce((out, a) => "
        "{ out[a] = Number(r.touched[`Model!${a}`]); return out }, {});"
    )
    proc = run([*WITAN_CMD, "xlsx", "exec", str(dst), "--code", code, "--save"])
    return json.loads(proc.stdout)


def run_openpyxl() -> Path:
    dst = OUT / "case18_openpyxl.xlsx"
    shutil.copy(FIXTURE, dst)
    wb = load_workbook(dst)
    wb["Model"]["B4"] = 6000
    wb.save(dst)
    return dst


def main() -> int:
    shutil.rmtree(OUT, ignore_errors=True)
    OUT.mkdir(parents=True)

    witan = run_witan()
    openpyxl_path = run_openpyxl()
    openpyxl_cached = read_cells(openpyxl_path, data_only=True)
    lo_path = resave_with_libreoffice(openpyxl_path)
    lo_cached = read_cells(lo_path, data_only=True)
    lo_formulas = read_cells(lo_path, data_only=False)

    witan_ok = witan == EXPECTED
    openpyxl_ok = openpyxl_cached == EXPECTED
    lo_ok = lo_cached == EXPECTED
    lo_failed_as_expected = lo_cached["D1"] == 14000 and all(lo_cached[a] is None for a in KEY_CELLS if a != "D1")

    print("18 What-If Data Table recalculation")
    print(f"   expected: {EXPECTED}")
    print(f"   witan: {witan}")
    print(f"   openpyxl data_only after write: {openpyxl_cached}")
    print(f"   openpyxl + LibreOffice data_only: {lo_cached}")
    print(f"   formulas after LibreOffice: {lo_formulas}")
    print()

    unexpected = not witan_ok or openpyxl_ok or lo_ok or not lo_failed_as_expected
    print(f"Summary: {'expected comparison' if not unexpected else 'unexpected result'}")
    return 1 if unexpected else 0


if __name__ == "__main__":
    raise SystemExit(main())
