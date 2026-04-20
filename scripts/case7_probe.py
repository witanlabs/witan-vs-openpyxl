"""Probe: what does openpyxl expose about the 2-variable data table in sensitivity2d.xlsx?"""
import sys
from openpyxl import load_workbook

PATH = sys.argv[1]
wb = load_workbook(PATH)
ws = wb["Model"]

print("=== Public data-table-ish attributes on Worksheet ===")
for name in sorted(dir(ws)):
    if "table" in name.lower() or "data" in name.lower() or "pivot" in name.lower():
        print(f"  ws.{name}")

print()
print("=== ws.tables (ListObjects) ===")
print(f"  count = {len(ws.tables)}")
for k, v in ws.tables.items():
    print(f"  {k!r} -> {v}")

print()
print("=== Private data_table attrs ===")
for name in ["_data_tables", "data_tables", "_dataTables", "dataTables"]:
    if hasattr(ws, name):
        v = getattr(ws, name)
        print(f"  ws.{name} = {v!r}")
    else:
        print(f"  ws.{name}: not found")

print()
print("=== Inspect formula cells in the data-table region ===")
for row in range(1, 7):
    for col_letter in "DEFGHI":
        addr = f"{col_letter}{row}"
        c = ws[addr]
        print(f"  {addr}: value={c.value!r}")
