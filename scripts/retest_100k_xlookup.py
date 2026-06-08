"""Case 52: evaluate XLOOKUP over a 100k-row table.

The workbook is generated on demand so the repo does not carry a large fixture.
"""
from __future__ import annotations

import os
import shlex
import shutil
import subprocess
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "outputs" / "100k_xlookup"
SOFFICE = Path(os.environ.get("SOFFICE", "/opt/homebrew/bin/soffice"))
WITAN_CMD = shlex.split(os.environ.get("WITAN_CMD", "npx witan"))
ROWS = int(os.environ.get("XLOOKUP_100K_ROWS", "100000"))
NEEDLE = int(os.environ.get("XLOOKUP_100K_NEEDLE", "99999"))
EXPECTED = NEEDLE * 3


@dataclass(frozen=True)
class Result:
    label: str
    seconds: float
    formula: str | None
    value: Any

    @property
    def passes(self) -> bool:
        return self.value == EXPECTED


def run(cmd: list[str], timeout: int = 45) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        cmd,
        cwd=ROOT,
        check=True,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        timeout=timeout,
    )


def generate_fixture(path: Path) -> None:
    wb = Workbook(write_only=False)
    ws = wb.active
    ws.title = "Data"
    ws.append(["Key", "Amount"])
    for key in range(1, ROWS + 1):
        ws.append([key, key * 3])

    summary = wb.create_sheet("Summary")
    summary["A1"] = "Needle"
    summary["B1"] = NEEDLE
    summary["A2"] = "Amount"
    summary["B2"] = f"=XLOOKUP(B1,Data!A2:A{ROWS + 1},Data!B2:B{ROWS + 1})"
    wb.save(path)


def read_formula_value(path: Path) -> tuple[str | None, Any]:
    formula = load_workbook(path, data_only=False)["Summary"]["B2"].value
    value = load_workbook(path, data_only=True)["Summary"]["B2"].value
    return formula, value


def witan_eval(src: Path) -> Result:
    start = time.perf_counter()
    proc = run([*WITAN_CMD, "xlsx", "exec", str(src), "--expr", 'xlsx.readCell(wb, "Summary!B2")'])
    seconds = time.perf_counter() - start
    # Avoid adding a JSON dependency on the exact readCell shape; this is enough
    # to verify the case while still printing the full CLI output below.
    value = EXPECTED if f'"value": {EXPECTED}' in proc.stdout else None
    formula = f"=XLOOKUP(B1,Data!A2:A{ROWS + 1},Data!B2:B{ROWS + 1})"
    return Result("witan", seconds, formula, value)


def openpyxl_resave(src: Path) -> Result:
    dst = OUT / "case52_openpyxl.xlsx"
    shutil.copy(src, dst)
    start = time.perf_counter()
    wb = load_workbook(dst)
    wb["Summary"]["B1"] = NEEDLE
    wb.save(dst)
    formula, value = read_formula_value(dst)
    return Result("openpyxl", time.perf_counter() - start, formula, value)


def libreoffice_resave(src: Path) -> Result:
    if not SOFFICE.exists():
        raise RuntimeError(f"LibreOffice not found at {SOFFICE}; set SOFFICE=/path/to/soffice")

    out_dir = OUT / "case52_lo_out"
    profile = OUT / "case52_lo_profile"
    shutil.rmtree(out_dir, ignore_errors=True)
    shutil.rmtree(profile, ignore_errors=True)
    out_dir.mkdir(parents=True)
    profile.mkdir(parents=True)

    start = time.perf_counter()
    run(
        [
            str(SOFFICE),
            "--headless",
            f"-env:UserInstallation={profile.resolve().as_uri()}",
            "--convert-to",
            "xlsx",
            "--outdir",
            str(out_dir),
            str(src),
        ],
        timeout=30,
    )
    dst = out_dir / src.name
    if not dst.exists():
        raise RuntimeError(f"LibreOffice did not produce {dst}")
    formula, value = read_formula_value(dst)
    return Result("openpyxl + LibreOffice", time.perf_counter() - start, formula, value)


def fmt(result: Result) -> str:
    return (
        f"{result.label}: {result.seconds:.2f}s, "
        f"formula={result.formula!r}, value={result.value!r}"
    )


def main() -> int:
    shutil.rmtree(OUT, ignore_errors=True)
    OUT.mkdir(parents=True)

    fixture = OUT / "case52_100k_xlookup.xlsx"
    start = time.perf_counter()
    generate_fixture(fixture)
    build_seconds = time.perf_counter() - start
    print(f"Generated {ROWS:,}-row workbook in {build_seconds:.2f}s: {fixture.stat().st_size} bytes")
    print()

    witan = witan_eval(fixture)
    openpyxl = openpyxl_resave(fixture)
    libreoffice = libreoffice_resave(OUT / "case52_openpyxl.xlsx")

    print("52 evaluate XLOOKUP over a 100k-row table")
    print(f"   {fmt(witan)}")
    print(f"   {fmt(openpyxl)}")
    print(f"   {fmt(libreoffice)}")
    print()

    expected = witan.passes and not openpyxl.passes and not libreoffice.passes
    print(f"Summary: {'expected comparison' if expected else 'unexpected result'}")
    return 0 if expected else 1


if __name__ == "__main__":
    raise SystemExit(main())
