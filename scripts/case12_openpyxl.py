"""Reproduce openpyxl tracker issue: CellRichText with bold TextBlocks
separated only by whitespace produces a file Excel flags as corrupt."""
import sys
from openpyxl import Workbook
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText

OUT = sys.argv[1]
wb = Workbook()
ws = wb.active
ws.title = "Data"

bold = InlineFont(b=True)
strike_red = InlineFont(strike=True, color="00990000")

ws["A1"] = CellRichText(
    "Normal text ",
    TextBlock(bold, "Bold1"),
    " ",                        # single space between two bold runs
    TextBlock(bold, "Bold2"),
    " more normal",
)

ws["A2"] = CellRichText(
    "Leading ",
    TextBlock(bold, "Bold1"),
    "   ",                      # multiple spaces
    TextBlock(bold, "Bold2"),
    " trailing",
)

# The exact case from the issue: a lone whitespace-only TextBlock for a diff-style presentation
ws["A3"] = CellRichText(
    "Some text",
    TextBlock(strike_red, " "),    # whitespace-only TextBlock with formatting
    "and some more.",
)

wb.save(OUT)
print(f"saved {OUT}")

# Inspect the raw XML for xml:space="preserve" on whitespace-only runs
import zipfile
with zipfile.ZipFile(OUT) as z:
    xml = z.read("xl/worksheets/sheet1.xml").decode()
# Count <t> elements that contain only whitespace, with vs without xml:space="preserve"
import re
runs = re.findall(r'<t(\s+xml:space="preserve")?[^>]*>([^<]*)</t>', xml)
print()
print("=== Run analysis ===")
for attr, body in runs:
    if body and body.strip() == "" and len(body) > 0:
        print(f"  whitespace-only run: preserve-attr={bool(attr)!r}  body={body!r}")
    elif body.startswith(" ") or body.endswith(" "):
        print(f"  leading/trailing-space run: preserve-attr={bool(attr)!r}  body={body!r}")
