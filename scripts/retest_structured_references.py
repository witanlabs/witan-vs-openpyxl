"""Case 20: recalculate formulas that use Excel structured references.

The workbook contains a ListObject named SalesTbl and Summary formulas that
refer to SalesTbl[Amount] and SalesTbl[Region]. The task changes one amount
inside the table and reports the structured-reference formula results.
"""
from __future__ import annotations

import json
import os
import shlex
import shutil
import subprocess
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "outputs" / "structured_references"
SOFFICE = Path(os.environ.get("SOFFICE", "/opt/homebrew/bin/soffice"))
WITAN_CMD = shlex.split(os.environ.get("WITAN_CMD", "npx witan"))
EXPECTED = {"B1": 650, "B2": 400}


def run(cmd: list[str]) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        cmd,
        cwd=ROOT,
        check=True,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
    )


def build_fixture(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"
    rows = [
        ["Region", "Product", "Amount"],
        ["North", "A", 100],
        ["South", "B", 200],
        ["North", "C", 300],
    ]
    for row in rows:
        ws.append(row)

    table = Table(displayName="SalesTbl", ref="A1:C4")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    summary = wb.create_sheet("Summary")
    summary["A1"] = "Total amount"
    summary["B1"] = "=SUM(SalesTbl[Amount])"
    summary["A2"] = "North amount"
    summary["B2"] = '=SUMIFS(SalesTbl[Amount],SalesTbl[Region],"North")'
    wb.save(path)


def resave_with_libreoffice(src: Path) -> Path:
    if not SOFFICE.exists():
        raise RuntimeError(f"LibreOffice not found at {SOFFICE}; set SOFFICE=/path/to/soffice")

    work = OUT / "_lo_work"
    converted = OUT / "_lo_out"
    profile = OUT / "_lo_profile"
    shutil.rmtree(work, ignore_errors=True)
    shutil.rmtree(converted, ignore_errors=True)
    shutil.rmtree(profile, ignore_errors=True)
    work.mkdir(parents=True)
    converted.mkdir(parents=True)
    profile.mkdir(parents=True)

    local = work / src.name
    shutil.copy(src, local)
    run(
        [
            str(SOFFICE),
            "--headless",
            f"-env:UserInstallation={profile.resolve().as_uri()}",
            "--convert-to",
            "xlsx",
            "--outdir",
            str(converted),
            str(local),
        ]
    )
    out = converted / local.name
    if not out.exists():
        raise RuntimeError(f"LibreOffice did not produce {out}")
    return out


def run_witan(fixture: Path) -> dict[str, object]:
    dst = OUT / "case20_witan.xlsx"
    shutil.copy(fixture, dst)
    code = (
        'const r = await xlsx.setCells(wb, [{address:"Sales!C3", value:250}]); '
        'return {'
        '  touched:r.touched,'
        '  B1:(await xlsx.readCell(wb, "Summary!B1")).value,'
        '  B2:(await xlsx.readCell(wb, "Summary!B2")).value'
        '};'
    )
    proc = run([*WITAN_CMD, "xlsx", "exec", str(dst), "--code", code, "--save"])
    result = json.loads(proc.stdout)
    return {"B1": result["B1"], "B2": result["B2"]}


def run_openpyxl(fixture: Path) -> Path:
    dst = OUT / "case20_openpyxl.xlsx"
    shutil.copy(fixture, dst)
    wb = load_workbook(dst)
    wb["Sales"]["C3"] = 250
    wb.save(dst)
    return dst


def summary_values(path: Path, data_only: bool) -> dict[str, object]:
    ws = load_workbook(path, data_only=data_only)["Summary"]
    return {"B1": ws["B1"].value, "B2": ws["B2"].value}


def table_ref(path: Path) -> str:
    ws = load_workbook(path, data_only=False)["Sales"]
    return ws.tables["SalesTbl"].ref


def main() -> int:
    shutil.rmtree(OUT, ignore_errors=True)
    OUT.mkdir(parents=True)
    fixture = OUT / "case20_fixture.xlsx"
    build_fixture(fixture)

    witan = run_witan(fixture)
    openpyxl_path = run_openpyxl(fixture)
    openpyxl_cached = summary_values(openpyxl_path, data_only=True)
    lo_path = resave_with_libreoffice(openpyxl_path)
    lo_cached = summary_values(lo_path, data_only=True)
    lo_formulas = summary_values(lo_path, data_only=False)
    lo_table_ref = table_ref(lo_path)

    print("20 Structured-reference formula recalculation")
    print(f"   expected: {EXPECTED}")
    print(f"   witan: {witan}")
    print(f"   openpyxl data_only after write: {openpyxl_cached}")
    print(f"   openpyxl + LibreOffice data_only: {lo_cached}")
    print(f"   formulas after LibreOffice: {lo_formulas}")
    print(f"   table ref after LibreOffice: {lo_table_ref}")
    print()

    unexpected = witan != EXPECTED or openpyxl_cached == EXPECTED or lo_cached != EXPECTED or lo_table_ref != "A1:C4"
    print(f"Summary: {'expected comparison' if not unexpected else 'unexpected result'}")
    return 1 if unexpected else 0


if __name__ == "__main__":
    raise SystemExit(main())
