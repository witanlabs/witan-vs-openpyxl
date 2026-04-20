"""Task: add a resolved threaded comment on Data!B3 by 'Auditor' saying 'Verified against ledger',
then list every threaded comment with author, text, resolved state."""
import sys, shutil, zipfile
from openpyxl import load_workbook
from openpyxl.comments import Comment

SRC, DST = sys.argv[1], sys.argv[2]
shutil.copy(SRC, DST)

wb = load_workbook(DST)
ws = wb["Data"]

print("=== What openpyxl sees BEFORE writing ===")
for addr in ["B2", "C2", "B3"]:
    c = ws[addr]
    print(f"  {addr}.comment = {c.comment!r}")

# Best effort: add a comment (openpyxl only supports legacy Comment).
ws["B3"].comment = Comment("Verified against ledger", "Auditor")
wb.save(DST)

print("\n=== Parts in saved file ===")
with zipfile.ZipFile(DST) as z:
    for n in z.namelist():
        if "thread" in n.lower() or "person" in n.lower() or "comment" in n.lower():
            print(f"  {n}")

print("\n=== Re-read after save ===")
wb2 = load_workbook(DST)
ws2 = wb2["Data"]
for addr in ["B2", "C2", "B3"]:
    c = ws2[addr].comment
    print(f"  {addr}.comment = {c!r}")
