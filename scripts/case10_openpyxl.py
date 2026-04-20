"""Task: rename sheet "Inputs" to "Parameters". Report the resulting formulas
and cached values."""
import sys, shutil
from openpyxl import load_workbook

SRC, DST = sys.argv[1], sys.argv[2]
shutil.copy(SRC, DST)

wb = load_workbook(DST)
wb["Inputs"].title = "Parameters"
wb.save(DST)

print("=== After rename (formula view) ===")
wb2 = load_workbook(DST)
for addr in ("B1", "B2", "B3", "B4"):
    c = wb2["Summary"][addr]
    print(f"  Summary!{addr}  = {c.value!r}")

print()
print("=== Defined names ===")
for dn in wb2.defined_names.values() if hasattr(wb2.defined_names, 'values') else []:
    print(f"  {dn.name} -> {dn.attr_text!r}")
# alternate API
for name in wb2.defined_names:
    val = wb2.defined_names[name]
    print(f"  [name={name}] {val.attr_text!r}")

print()
print("=== Cached values (data_only) ===")
wb3 = load_workbook(DST, data_only=True)
for addr in ("B1", "B2", "B3", "B4", "B5", "B6"):
    print(f"  Summary!{addr}  = {wb3['Summary'][addr].value!r}")
