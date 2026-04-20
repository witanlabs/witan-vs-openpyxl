"""Task A — describe the data table: what inputs vary, what output is computed?
   Task B — change the row input values from [500,750,1000,1250,1500] to [600,800,1000,1200,1400]."""
import sys, shutil
from openpyxl import load_workbook

SRC, DST = sys.argv[1], sys.argv[2]
shutil.copy(SRC, DST)

wb = load_workbook(DST)
ws = wb["Model"]

print("=== Task A: describe the data table ===")
# openpyxl has no data-table API. Best the agent can do is probe ws.tables.
print(f"  ws.tables count (ListObjects only) = {len(ws.tables)}")
# Check for data-table attributes
for name in ("data_tables", "_data_tables", "dataTables", "_dataTables"):
    print(f"  ws.{name} exists: {hasattr(ws, name)}")

# Inspect the array-formula placeholder the agent can find via cell iteration
target = ws["E2"]
print(f"  ws['E2'].value type = {type(target.value).__name__}")
print(f"  ws['E2'].value repr = {target.value!r}")
# Try common DataTableFormula attributes
if hasattr(target.value, "__dict__"):
    print(f"  ws['E2'].value attrs = {vars(target.value)}")

print()
print("=== Task B: change row input values (E1:I1) ===")
# Naive approach: overwrite the displayed input values in E1:I1.
new_vals = [600, 800, 1000, 1200, 1400]
for col_letter, v in zip("EFGHI", new_vals):
    ws[f"{col_letter}1"] = v
wb.save(DST)

print("  wrote E1..I1 =", new_vals)
print("  saved", DST)

print()
print("=== Re-read to check data_only cache ===")
wb2 = load_workbook(DST, data_only=True)
ws2 = wb2["Model"]
for col_letter in "EFGHI":
    print(f"  {col_letter}1 = {ws2[f'{col_letter}1'].value!r}")
for r in range(2, 7):
    row_vals = [ws2[f"{c}{r}"].value for c in "EFGHI"]
    print(f"  row {r}: {row_vals}")
