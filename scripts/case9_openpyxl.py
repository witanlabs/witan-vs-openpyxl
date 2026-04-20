"""Per openpyxl tracker: Conditional Formatting Rule with MultiCellRange dies on save.

Two documented user paths for a CF rule over a discontiguous range:
  1. A comma-separated string ("A1:B1,A2:B2") — raises TypeError up-front
  2. A MultiCellRange object (the type the TypeError points at) — crashes at save

Space-separated ("A1:A10 D1:D10") happens to work, but that is not what the
openpyxl docs or the TypeError message suggest."""
import sys, openpyxl
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
from openpyxl.worksheet.cell_range import MultiCellRange, CellRange

OUT = sys.argv[1]
rule = CellIsRule(
    operator="greaterThan", formula=["100"],
    fill=PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
)

def seed(ws):
    import random; random.seed(0)
    for r in range(1, 11):
        ws.cell(row=r, column=1, value=random.randint(0, 200))
        ws.cell(row=r, column=4, value=random.randint(0, 200))

# --- Path 1: comma-separated string, what a user would try first ---
print("=== Path 1: add_conditional_formatting(\"A1:B1,A2:B2\", rule) ===")
wb1 = openpyxl.Workbook(); ws1 = wb1.active; ws1.title = "Data"; seed(ws1)
try:
    ws1.conditional_formatting.add("A1:A10,D1:D10", rule)
    wb1.save(OUT + ".commas.xlsx")
    print(f"  save OK")
except Exception as e:
    print(f"  FAILED ({type(e).__name__}): {e}")

# --- Path 2: explicit MultiCellRange (as TypeError suggests) ---
print()
print("=== Path 2: MultiCellRange([CellRange('A1:A10'), CellRange('D1:D10')]) ===")
wb2 = openpyxl.Workbook(); ws2 = wb2.active; ws2.title = "Data"; seed(ws2)
mcr = MultiCellRange([CellRange(range_string="A1:A10"), CellRange(range_string="D1:D10")])
try:
    ws2.conditional_formatting.add(mcr, rule)
    wb2.save(OUT)
    print(f"  save OK -> {OUT}")
except Exception as e:
    print(f"  FAILED ({type(e).__name__}): {e}")

# --- Path 3: space-separated (works, included as reference) ---
print()
print("=== Path 3 (reference): add_conditional_formatting(\"A1:A10 D1:D10\", rule) ===")
wb3 = openpyxl.Workbook(); ws3 = wb3.active; ws3.title = "Data"; seed(ws3)
try:
    ws3.conditional_formatting.add("A1:A10 D1:D10", rule)
    wb3.save(OUT + ".spaces.xlsx")
    print(f"  save OK")
except Exception as e:
    print(f"  FAILED ({type(e).__name__}): {e}")
