"""Cases 36-37: Witan render/preview fidelity versus LibreOffice PDF export.

The script generates two visual comparison cases and writes stable PNG assets
for the README:

- Case 36: a waterfall chart. Witan renders the intended waterfall; LibreOffice
  PDF export renders the same workbook as a malformed line-like chart.
- Case 37: a text-layout stress sheet. Witan renders the requested worksheet
  range directly; LibreOffice PDF export changes typography, scaling, and some
  text layout behavior.

The comparison is intentionally visual. openpyxl can create workbooks but has
no native render/preview operation. LibreOffice can export a PDF/PNG, but these
cases show it is not an equivalent substitute for Witan's sheet-range render.
"""
from __future__ import annotations

import os
import shutil
import subprocess
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "outputs" / "preview_rendering"
ASSETS = ROOT / "assets" / "preview_rendering"
SOFFICE = Path(os.environ.get("SOFFICE", "/opt/homebrew/bin/soffice"))
PDFTOPPM = Path(os.environ.get("PDFTOPPM", "/opt/homebrew/bin/pdftoppm"))
XLSX_SERVE = Path(
    os.environ.get(
        "XLSX_SERVE",
        str((ROOT / "../witan-alfred/bin/publish/xlsx-serve").resolve()),
    )
)


def run(cmd: list[str]) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        cmd,
        cwd=ROOT,
        check=True,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
    )


def libreoffice_pdf_to_png(src: Path, dst: Path) -> None:
    if not SOFFICE.exists():
        raise RuntimeError(f"LibreOffice not found at {SOFFICE}; set SOFFICE=/path/to/soffice")
    if not PDFTOPPM.exists():
        raise RuntimeError(f"pdftoppm not found at {PDFTOPPM}; set PDFTOPPM=/path/to/pdftoppm")

    converted = OUT / "_lo_pdf"
    profile = OUT / "_lo_profile"
    shutil.rmtree(converted, ignore_errors=True)
    shutil.rmtree(profile, ignore_errors=True)
    converted.mkdir(parents=True)
    profile.mkdir(parents=True)

    run(
        [
            str(SOFFICE),
            "--headless",
            f"-env:UserInstallation={profile.resolve().as_uri()}",
            "--convert-to",
            "pdf",
            "--outdir",
            str(converted),
            str(src),
        ]
    )
    pdf = converted / f"{src.stem}.pdf"
    if not pdf.exists():
        raise RuntimeError(f"LibreOffice did not produce {pdf}")

    tmp_prefix = dst.with_suffix("")
    run([str(PDFTOPPM), "-png", "-singlefile", "-r", "110", str(pdf), str(tmp_prefix)])
    produced = tmp_prefix.with_suffix(".png")
    if produced != dst:
        produced.replace(dst)


def render_with_witan(src: Path, range_address: str, dst: Path) -> None:
    run([str(XLSX_SERVE), "render", str(src), "-r", range_address, "-o", str(dst), "--dpr", "1"])
    if not dst.exists():
        raise RuntimeError(f"Witan render did not produce {dst}")


def build_waterfall_workbook(dst: Path) -> None:
    code = r'''
await xlsx.addSheet(wb, "Summary");
await xlsx.setCells(wb, [
  {address:"Summary!A1", value:"Step"},       {address:"Summary!B1", value:"Amount"},
  {address:"Summary!A2", value:"Revenue"},    {address:"Summary!B2", value:1200},
  {address:"Summary!A3", value:"COGS"},       {address:"Summary!B3", value:-420},
  {address:"Summary!A4", value:"Payroll"},    {address:"Summary!B4", value:-250},
  {address:"Summary!A5", value:"Gross"},      {address:"Summary!B5", value:530},
  {address:"Summary!A6", value:"Tax"},        {address:"Summary!B6", value:-95},
  {address:"Summary!A7", value:"Net"},        {address:"Summary!B7", value:435}
]);
await xlsx.addChart(wb, "Summary", {
  name: "Profit Bridge",
  position: { from: { cell: "D2" }, to: { cell: "L18" } },
  groups: [{
    type: "waterfall",
    series: [{
      name: { ref: "Summary!B1" },
      categories: "Summary!A2:A7",
      values: "Summary!B2:B7",
      totalIndexes: [3, 5],
      dataLabels: { showValue: true, position: "outsideEnd" }
    }]
  }],
  title: { text: "Profit Bridge" },
  legend: { position: "topRight" },
  axes: { value: { numberFormat: "$#,##0" } },
  styleId: 395
});
return true;
'''
    run([str(XLSX_SERVE), "exec", str(dst), "--create", "--save", "--code", code])


def build_text_workbook(dst: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Layout"
    ws.sheet_view.showGridLines = False

    for col, width in [
        ("A", 16),
        ("B", 10),
        ("C", 12),
        ("D", 16),
        ("E", 12),
        ("F", 10),
        ("G", 14),
        ("H", 14),
    ]:
        ws.column_dimensions[col].width = width
    for row, height in [(1, 28), (2, 54), (3, 42), (4, 38), (5, 60), (6, 26), (7, 34), (8, 34), (9, 34), (10, 34)]:
        ws.row_dimensions[row].height = height

    ws["A1"] = "Text layout stress test"
    ws["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="274060")
    ws.merge_cells("A1:H1")

    ws["A2"] = "Wrapped text that should occupy exactly three lines without clipping in a narrow cell"
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["B2"] = "Rotated 45 degrees"
    ws["B2"].alignment = Alignment(textRotation=45, horizontal="center", vertical="center")
    ws["C2"] = "Shrink to fit: a very long single line"
    ws["C2"].alignment = Alignment(shrink_to_fit=True, horizontal="center")
    ws["D2"] = "Centered across merged cells with border"
    ws.merge_cells("D2:F3")
    ws["D2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["D2"].font = Font(bold=True)
    ws["G2"] = "Indented\nmanual breaks"
    ws["G2"].alignment = Alignment(wrap_text=True, indent=2, vertical="center")
    ws["A4"] = "Tall wrapped text with row height that may clip descenders and final line"
    ws["A4"].alignment = Alignment(wrap_text=True, vertical="center")
    ws["B4"] = "90deg"
    ws["B4"].alignment = Alignment(textRotation=90, horizontal="center", vertical="center")
    ws["C4"] = "Stacked-ish text"
    ws["C4"].alignment = Alignment(textRotation=255, horizontal="center", vertical="center")

    ws["D5"] = "Currency"
    ws["E5"] = "Percent"
    ws["F5"] = "Date"
    ws["G5"] = "Long code"
    ws["D6"] = 1234.5
    ws["D6"].number_format = "$#,##0.00"
    ws["E6"] = 0.1234
    ws["E6"].number_format = "0.00%"
    ws["F6"] = "2026-06-03"
    ws["G6"] = "ABCDEFGHIJKLMN"

    thin = Side(style="thin", color="BFBFBF")
    for row in ws.iter_rows(min_row=1, max_row=10, min_col=1, max_col=8):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb.save(dst)


def main() -> int:
    shutil.rmtree(OUT, ignore_errors=True)
    OUT.mkdir(parents=True)
    ASSETS.mkdir(parents=True, exist_ok=True)

    waterfall = OUT / "case36_waterfall.xlsx"
    build_waterfall_workbook(waterfall)
    render_with_witan(waterfall, "Summary!A1:L20", ASSETS / "case36_witan_waterfall.png")
    libreoffice_pdf_to_png(waterfall, ASSETS / "case36_libreoffice_waterfall.png")

    text = OUT / "case37_text_layout.xlsx"
    build_text_workbook(text)
    render_with_witan(text, "Layout!A1:H10", ASSETS / "case37_witan_text_layout.png")
    libreoffice_pdf_to_png(text, ASSETS / "case37_libreoffice_text_layout.png")

    print("36 Preview waterfall chart rendering")
    print(f"   witan: {ASSETS / 'case36_witan_waterfall.png'}")
    print(f"   LibreOffice PDF export: {ASSETS / 'case36_libreoffice_waterfall.png'}")
    print()
    print("37 Preview text layout rendering")
    print(f"   witan: {ASSETS / 'case37_witan_text_layout.png'}")
    print(f"   LibreOffice PDF export: {ASSETS / 'case37_libreoffice_text_layout.png'}")
    print()
    print("Summary: generated preview comparison images")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
