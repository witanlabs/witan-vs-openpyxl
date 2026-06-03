"""Cases 31-35: Witan lint diagnostics with no openpyxl/LibreOffice equivalent.

These are verifier cases rather than authoring cases. Witan reports workbook
risks through `witan xlsx lint`; openpyxl has no comparable semantic lint API,
and passive LibreOffice open/save does not report these diagnostics.
"""
from __future__ import annotations

import json
import os
import shutil
import subprocess
from pathlib import Path
from typing import Any, Callable

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "outputs" / "lint_cases"
SOFFICE = Path(os.environ.get("SOFFICE", "/opt/homebrew/bin/soffice"))


def run(cmd: list[str], check: bool = True) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        cmd,
        cwd=ROOT,
        check=check,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
    )


def resave_with_openpyxl(src: Path, dst: Path) -> str:
    wb = load_workbook(src)
    wb.save(dst)
    return "saved without semantic diagnostics"


def resave_with_libreoffice(src: Path) -> str:
    if not SOFFICE.exists():
        raise RuntimeError(f"LibreOffice not found at {SOFFICE}; set SOFFICE=/path/to/soffice")

    work = OUT / "_lo_work"
    converted = OUT / "_lo_out"
    profile = OUT / "_lo_profile"
    shutil.rmtree(work, ignore_errors=True)
    shutil.rmtree(converted, ignore_errors=True)
    shutil.rmtree(profile, ignore_errors=True)
    work.mkdir(parents=True)
    converted.mkdir(parents=True)
    profile.mkdir(parents=True)

    local = work / src.name
    shutil.copy(src, local)
    proc = run(
        [
            str(SOFFICE),
            "--headless",
            f"-env:UserInstallation={profile.resolve().as_uri()}",
            "--convert-to",
            "xlsx",
            "--outdir",
            str(converted),
            str(local),
        ]
    )
    if not (converted / local.name).exists():
        raise RuntimeError(f"LibreOffice did not produce {converted / local.name}")
    return "saved without semantic diagnostics" if "D00" not in proc.stdout else proc.stdout


def witan_lint(path: Path) -> list[dict[str, Any]]:
    proc = run(["npx", "witan", "xlsx", "lint", str(path), "--json"], check=False)
    payload = json.loads(proc.stdout)
    return payload.get("diagnostics", [])


def make_unsorted_lookup(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for row in [["Key", "Value"], [1, "one"], [3, "three"], [2, "two"]]:
        ws.append(row)
    ws["D1"] = "=VLOOKUP(2,A2:B4,2,TRUE)"
    wb.save(path)


def make_duplicate_lookup(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for row in [["Key", "Value"], ["ID001", "first"], ["ID001", "second"], ["ID002", "third"]]:
        ws.append(row)
    ws["D1"] = '=VLOOKUP("ID001",A2:B4,2,FALSE)'
    wb.save(path)


def make_empty_coercion(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "=B1+5"
    wb.save(path)


def make_numeric_ignores_text(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = 1
    ws["A2"] = "oops"
    ws["A3"] = 3
    ws["B1"] = "=SUM(A1:A3)"
    wb.save(path)


def make_mixed_currency(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = 10
    ws["A1"].number_format = "$#,##0.00"
    ws["A2"] = 20
    ws["A2"].number_format = "€#,##0.00"
    ws["A3"] = "=SUM(A1:A2)"
    wb.save(path)


CASES: list[tuple[int, str, str, Callable[[Path], None]]] = [
    (31, "D002", "Approximate lookup over an unsorted range", make_unsorted_lookup),
    (32, "D007", "Duplicate lookup keys", make_duplicate_lookup),
    (33, "D003", "Empty cell coerced in arithmetic", make_empty_coercion),
    (34, "D005", "SUM silently ignores text in a numeric range", make_numeric_ignores_text),
    (35, "D008", "Mixed currencies in an aggregate", make_mixed_currency),
]


def summarize(diags: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "ruleId": diag.get("ruleId"),
            "severity": diag.get("severity"),
            "location": diag.get("location"),
            "message": diag.get("message"),
        }
        for diag in diags
    ]


def main() -> int:
    shutil.rmtree(OUT, ignore_errors=True)
    OUT.mkdir(parents=True)

    unexpected = False
    for number, rule_id, title, make in CASES:
        src = OUT / f"case{number}_{rule_id.lower()}.xlsx"
        openpyxl_out = OUT / f"case{number}_{rule_id.lower()}_openpyxl.xlsx"
        make(src)

        diags = summarize(witan_lint(src))
        openpyxl_result = resave_with_openpyxl(src, openpyxl_out)
        libreoffice_result = resave_with_libreoffice(openpyxl_out)

        print(f"{number} Lint {rule_id}: {title}")
        print(f"   witan lint: {diags}")
        print(f"   openpyxl: {openpyxl_result}")
        print(f"   openpyxl + LibreOffice: {libreoffice_result}")
        print()

        unexpected = unexpected or not any(diag["ruleId"] == rule_id for diag in diags)

    print(f"Summary: {'expected diagnostics' if not unexpected else 'unexpected result'}")
    return 1 if unexpected else 0


if __name__ == "__main__":
    raise SystemExit(main())
