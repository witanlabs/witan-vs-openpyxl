"""Cases 26-28: operation features from the Witan CLI changelog.

These cases cover high-level workbook operations added before the chart
features:

- copyRange adjusts relative formulas when copying.
- sortRange actually reorders worksheet rows.
- autoFitColumns writes concrete column widths from rendered content.

openpyxl can serialize related OOXML hints or manually copy cell strings, but
it does not provide equivalent high-level operations. LibreOffice passive
open/recalculate/save does not infer the missing operation after the fact.
"""
from __future__ import annotations

import json
import os
import re
import shlex
import shutil
import subprocess
import zipfile
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "outputs" / "changelog_operations"
SOFFICE = Path(os.environ.get("SOFFICE", "/opt/homebrew/bin/soffice"))
WITAN_CMD = shlex.split(os.environ.get("WITAN_CMD", "npx witan"))


def run(cmd: list[str]) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        cmd,
        cwd=ROOT,
        check=True,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
    )


def resave_with_libreoffice(src: Path) -> Path:
    if not SOFFICE.exists():
        raise RuntimeError(f"LibreOffice not found at {SOFFICE}; set SOFFICE=/path/to/soffice")

    work = OUT / "_lo_work"
    converted = OUT / "_lo_out"
    profile = OUT / "_lo_profile"
    shutil.rmtree(work, ignore_errors=True)
    shutil.rmtree(converted, ignore_errors=True)
    shutil.rmtree(profile, ignore_errors=True)
    work.mkdir(parents=True)
    converted.mkdir(parents=True)
    profile.mkdir(parents=True)

    local = work / src.name
    shutil.copy(src, local)
    run(
        [
            str(SOFFICE),
            "--headless",
            f"-env:UserInstallation={profile.resolve().as_uri()}",
            "--convert-to",
            "xlsx",
            "--outdir",
            str(converted),
            str(local),
        ]
    )
    out = converted / local.name
    if not out.exists():
        raise RuntimeError(f"LibreOffice did not produce {out}")
    return out


def witan_create(path: Path, code: str) -> Any:
    proc = run([*WITAN_CMD, "xlsx", "exec", str(path), "--create", "--save", "--code", code])
    return json.loads(proc.stdout)


def read_formula_value(path: Path, address: str) -> dict[str, Any]:
    sheet, cell = address.split("!")
    formula = load_workbook(path, data_only=False)[sheet][cell].value
    value = load_workbook(path, data_only=True)[sheet][cell].value
    return {"formula": formula, "value": value}


def read_rows(path: Path) -> list[list[Any]]:
    ws = load_workbook(path, data_only=True)["Sheet1"]
    return [[ws.cell(row, col).value for col in (1, 2)] for row in range(1, 5)]


def column_a_width(path: Path) -> dict[str, Any]:
    wb = load_workbook(path)
    width = wb["Sheet1"].column_dimensions["A"].width
    with zipfile.ZipFile(path) as archive:
        xml = archive.read("xl/worksheets/sheet1.xml").decode("utf-8")
    match = re.search(r'<col[^>]*min="1"[^>]*/>', xml)
    return {"width": width, "xml": match.group(0) if match else None}


def witan_copy_range() -> dict[str, Any]:
    return witan_create(
        OUT / "case26_witan.xlsx",
        r'''
await xlsx.setCells(wb, [
  {address:"Sheet1!A1", value:10},
  {address:"Sheet1!A2", value:7},
  {address:"Sheet1!B1", formula:"=A1*2"}
]);
await xlsx.copyRange(wb, "Sheet1!B1", "Sheet1!B2");
return await xlsx.readCell(wb, "Sheet1!B2");
''',
    )


def openpyxl_copy_range() -> Path:
    dst = OUT / "case26_openpyxl.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = 10
    ws["A2"] = 7
    ws["B1"] = "=A1*2"
    ws["B2"] = ws["B1"].value
    wb.save(dst)
    return dst


def witan_sort_range() -> str:
    return witan_create(
        OUT / "case27_witan.xlsx",
        r'''
await xlsx.setCells(wb, [
  {address:"Sheet1!A1", value:"Name"},  {address:"Sheet1!B1", value:"Score"},
  {address:"Sheet1!A2", value:"Alpha"}, {address:"Sheet1!B2", value:1},
  {address:"Sheet1!A3", value:"Beta"},  {address:"Sheet1!B3", value:3},
  {address:"Sheet1!A4", value:"Gamma"}, {address:"Sheet1!B4", value:2}
]);
await xlsx.sortRange(wb, "Sheet1!A1:B4", [{col:"B", order:"desc"}]);
return await xlsx.readRangeTsv(wb, "Sheet1!A1:B4");
''',
    )


def openpyxl_sort_condition() -> Path:
    dst = OUT / "case27_openpyxl.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for row in [["Name", "Score"], ["Alpha", 1], ["Beta", 3], ["Gamma", 2]]:
        ws.append(row)
    ws.auto_filter.ref = "A1:B4"
    ws.auto_filter.add_sort_condition("B2:B4", descending=True)
    wb.save(dst)
    return dst


def witan_autofit() -> dict[str, Any]:
    return witan_create(
        OUT / "case28_witan.xlsx",
        r'''
await xlsx.setCells(wb, [
  {address:"Sheet1!A1", value:"This is a deliberately long heading that should force a wide column"},
  {address:"Sheet1!B1", value:"x"}
]);
return await xlsx.autoFitColumns(wb, "Sheet1", ["A", "B"], {minWidth:5, maxWidth:40});
''',
    )


def openpyxl_best_fit_hint() -> Path:
    dst = OUT / "case28_openpyxl.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "This is a deliberately long heading that should force a wide column"
    ws["B1"] = "x"
    ws.column_dimensions["A"].auto_size = True
    wb.save(dst)
    return dst


def main() -> int:
    shutil.rmtree(OUT, ignore_errors=True)
    OUT.mkdir(parents=True)

    copy_witan = witan_copy_range()
    copy_openpyxl_path = openpyxl_copy_range()
    copy_openpyxl = read_formula_value(copy_openpyxl_path, "Sheet1!B2")
    copy_lo = read_formula_value(resave_with_libreoffice(copy_openpyxl_path), "Sheet1!B2")

    sort_witan = witan_sort_range()
    sort_openpyxl_path = openpyxl_sort_condition()
    sort_openpyxl = read_rows(sort_openpyxl_path)
    sort_lo = read_rows(resave_with_libreoffice(sort_openpyxl_path))

    fit_witan = witan_autofit()
    fit_openpyxl_path = openpyxl_best_fit_hint()
    fit_openpyxl = column_a_width(fit_openpyxl_path)
    fit_lo = column_a_width(resave_with_libreoffice(fit_openpyxl_path))

    print("26 copyRange adjusts formulas")
    print(f"   witan: {{'formula': {copy_witan.get('formula')!r}, 'value': {copy_witan.get('value')!r}}}")
    print(f"   openpyxl: {copy_openpyxl}")
    print(f"   openpyxl + LibreOffice: {copy_lo}")
    print()

    print("27 sortRange reorders rows")
    print(f"   witan: {sort_witan!r}")
    print(f"   openpyxl: {sort_openpyxl}")
    print(f"   openpyxl + LibreOffice: {sort_lo}")
    print()

    print("28 autoFitColumns writes concrete widths")
    print(f"   witan: {fit_witan}")
    print(f"   openpyxl: {fit_openpyxl}")
    print(f"   openpyxl + LibreOffice: {fit_lo}")
    print()

    unexpected = (
        copy_witan.get("formula") != "=A2*2"
        or copy_witan.get("value") != 14
        or copy_lo.get("formula") == "=A2*2"
        or sort_witan != "A1|Name\tB1|Score\nA2|Beta\tB2|3\nA3|Gamma\tB3|2\nA4|Alpha\tB4|1"
        or sort_lo[1][0] == "Beta"
        or fit_witan.get("A", {}).get("width") != 40
        or fit_lo.get("width") == 40
    )
    print(f"Summary: {'expected comparisons' if not unexpected else 'unexpected result'}")
    return 1 if unexpected else 0


if __name__ == "__main__":
    raise SystemExit(main())
