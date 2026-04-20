"""Task B — extend the data table: add a new row for price=90 so the table
computes profit for prices 40..90 across the same volume columns."""
import sys, shutil
from openpyxl import load_workbook

SRC, DST = sys.argv[1], sys.argv[2]
shutil.copy(SRC, DST)

wb = load_workbook(DST)
ws = wb["Model"]

# Naive agent approach: write the new price label, let Excel extend.
ws["D7"] = 90
wb.save(DST)
print("  wrote D7 = 90")
print("  saved", DST)

print()
print("=== Re-read ===")
wb2 = load_workbook(DST, data_only=True)
ws2 = wb2["Model"]
print(f"  D7 = {ws2['D7'].value!r}")
for col_letter in "EFGHI":
    print(f"  {col_letter}7 = {ws2[f'{col_letter}7'].value!r}")

# Check that the DataTableFormula ref still says E2:I6
print()
print("=== DataTableFormula ref after save ===")
wb3 = load_workbook(DST)
dt = wb3["Model"]["E2"].value
if hasattr(dt, "__dict__"):
    print(f"  E2 ref/dt2D = {vars(dt)}")
