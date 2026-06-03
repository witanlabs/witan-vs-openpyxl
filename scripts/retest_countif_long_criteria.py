"""Case 30: COUNTIF with criteria longer than 255 characters.

Excel returns #VALUE! for COUNTIF-style criteria strings longer than 255
characters. Witan matches that behavior. LibreOffice 26.2.1.2 counts matching
cells instead, so the openpyxl + LibreOffice pairing reports a different value.
"""
from __future__ import annotations

import json
import os
import shlex
import shutil
import subprocess
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "outputs" / "countif_long_criteria"
SOFFICE = Path(os.environ.get("SOFFICE", "/opt/homebrew/bin/soffice"))
WITAN_CMD = shlex.split(os.environ.get("WITAN_CMD", "npx witan"))


def run(cmd: list[str]) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        cmd,
        cwd=ROOT,
        check=True,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
    )


def resave_with_libreoffice(src: Path) -> Path:
    if not SOFFICE.exists():
        raise RuntimeError(f"LibreOffice not found at {SOFFICE}; set SOFFICE=/path/to/soffice")

    work = OUT / "_lo_work"
    converted = OUT / "_lo_out"
    profile = OUT / "_lo_profile"
    shutil.rmtree(work, ignore_errors=True)
    shutil.rmtree(converted, ignore_errors=True)
    shutil.rmtree(profile, ignore_errors=True)
    work.mkdir(parents=True)
    converted.mkdir(parents=True)
    profile.mkdir(parents=True)

    local = work / src.name
    shutil.copy(src, local)
    run(
        [
            str(SOFFICE),
            "--headless",
            f"-env:UserInstallation={profile.resolve().as_uri()}",
            "--convert-to",
            "xlsx",
            "--outdir",
            str(converted),
            str(local),
        ]
    )
    out = converted / local.name
    if not out.exists():
        raise RuntimeError(f"LibreOffice did not produce {out}")
    return out


def witan_create() -> dict[str, Any]:
    dst = OUT / "case30_witan.xlsx"
    long_text = "x" * 260
    code = rf'''
const longText = "{long_text}";
await xlsx.setCells(wb, [
  {{address:"Sheet1!A1", value:longText}},
  {{address:"Sheet1!A2", value:longText}},
  {{address:"Sheet1!A3", value:longText.slice(0, -1) + "y"}},
  {{address:"Sheet1!B1", value:longText}},
  {{address:"Sheet1!D1", formula:"=COUNTIF(A1:A3,B1)"}}
]);
return await xlsx.readCell(wb, "Sheet1!D1");
'''
    proc = run([*WITAN_CMD, "xlsx", "exec", str(dst), "--create", "--save", "--code", code])
    cell = json.loads(proc.stdout)
    return {"formula": cell.get("formula"), "value": cell.get("value")}


def openpyxl_create() -> Path:
    dst = OUT / "case30_openpyxl.xlsx"
    long_text = "x" * 260
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = long_text
    ws["A2"] = long_text
    ws["A3"] = long_text[:-1] + "y"
    ws["B1"] = long_text
    ws["D1"] = "=COUNTIF(A1:A3,B1)"
    wb.save(dst)
    return dst


def read_cell(path: Path) -> dict[str, Any]:
    formula_wb = load_workbook(path, data_only=False)
    value_wb = load_workbook(path, data_only=True)
    return {
        "formula": formula_wb["Sheet1"]["D1"].value,
        "value": value_wb["Sheet1"]["D1"].value,
    }


def main() -> int:
    shutil.rmtree(OUT, ignore_errors=True)
    OUT.mkdir(parents=True)

    witan = witan_create()
    openpyxl_path = openpyxl_create()
    openpyxl = read_cell(openpyxl_path)
    libreoffice = read_cell(resave_with_libreoffice(openpyxl_path))

    print("30 COUNTIF long criteria")
    print(f"   witan: {witan}")
    print(f"   openpyxl: {openpyxl}")
    print(f"   openpyxl + LibreOffice: {libreoffice}")
    print()

    unexpected = witan["value"] != "#VALUE!" or libreoffice["value"] == "#VALUE!"
    print(f"Summary: {'expected comparison' if not unexpected else 'unexpected result'}")
    return 1 if unexpected else 0


if __name__ == "__main__":
    raise SystemExit(main())
