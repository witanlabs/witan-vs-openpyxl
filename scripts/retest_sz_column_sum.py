"""Case 53: sum column T in the sz-test workbook.

Set SZ_TEST_XLSX=/path/to/sz-test.xlsx to run against another copy.
"""
from __future__ import annotations

import os
import shlex
import shutil
import subprocess
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "outputs" / "sz_column_sum"
SOURCE = Path(os.environ.get("SZ_TEST_XLSX", str(ROOT / "fixtures" / "sz-test.xlsx")))
SOFFICE = Path(os.environ.get("SOFFICE", "/opt/homebrew/bin/soffice"))
WITAN_CMD = shlex.split(os.environ.get("WITAN_CMD", "npx witan"))
EXPECTED = Decimal("2058")


@dataclass(frozen=True)
class SumResult:
    label: str
    total: Decimal
    values: list[Any]
    non_numeric: list[tuple[int, Any]]

    @property
    def passes(self) -> bool:
        return self.total == EXPECTED and not self.non_numeric


def run(cmd: list[str], timeout: int = 30) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        cmd,
        cwd=ROOT,
        check=True,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        timeout=timeout,
    )


def openpyxl_sum(path: Path, label: str) -> SumResult:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    values: list[Any] = []
    total = Decimal("0")
    non_numeric: list[tuple[int, Any]] = []
    for row in range(1, ws.max_row + 1):
        value = ws[f"T{row}"].value
        values.append(value)
        if isinstance(value, (int, float, Decimal)):
            total += Decimal(str(value))
        elif value is not None:
            non_numeric.append((row, value))
    return SumResult(label, total, values, non_numeric)


def witan_sum(path: Path) -> SumResult:
    code = """
const data = await xlsx.readRange(wb, "Sheet1!T1:T30");
let sum = 0;
const values = [];
for (const row of data) {
  const cell = row[0];
  const value = cell && typeof cell === "object" && "value" in cell ? cell.value : cell;
  values.push(value);
  if (typeof value === "number") sum += value;
}
return {sum, values};
"""
    proc = run([*WITAN_CMD, "xlsx", "exec", str(path), "--code", code])
    # The output is JSON, but parsing loosely keeps the script independent of
    # any future readRange metadata shape.
    import json

    parsed = json.loads(proc.stdout)
    values = parsed["values"]
    non_numeric = [(idx + 1, value) for idx, value in enumerate(values) if not isinstance(value, (int, float))]
    return SumResult("witan", Decimal(str(parsed["sum"])), values, non_numeric)


def libreoffice_resave(path: Path) -> Path:
    if not SOFFICE.exists():
        raise RuntimeError(f"LibreOffice not found at {SOFFICE}; set SOFFICE=/path/to/soffice")

    out_dir = OUT / "lo_out"
    profile = OUT / "lo_profile"
    shutil.rmtree(out_dir, ignore_errors=True)
    shutil.rmtree(profile, ignore_errors=True)
    out_dir.mkdir(parents=True)
    profile.mkdir(parents=True)
    run(
        [
            str(SOFFICE),
            "--headless",
            f"-env:UserInstallation={profile.resolve().as_uri()}",
            "--convert-to",
            "xlsx",
            "--outdir",
            str(out_dir),
            str(path),
        ]
    )
    out = out_dir / path.name
    if not out.exists():
        raise RuntimeError(f"LibreOffice did not produce {out}")
    return out


def fmt(result: SumResult) -> str:
    preview = result.values[:10]
    suffix = "" if len(result.values) <= 10 else " ..."
    return (
        f"{result.label}: sum={result.total}, first values={preview}{suffix}, "
        f"non_numeric={result.non_numeric[:5]}"
    )


def main() -> int:
    if not SOURCE.exists():
        raise RuntimeError(f"Missing workbook: {SOURCE}")

    shutil.rmtree(OUT, ignore_errors=True)
    OUT.mkdir(parents=True)

    witan = witan_sum(SOURCE)
    openpyxl = openpyxl_sum(SOURCE, "openpyxl")
    libreoffice = openpyxl_sum(libreoffice_resave(SOURCE), "openpyxl + LibreOffice")

    print("53 sum column T in sz-test.xlsx")
    print(f"   {fmt(witan)}")
    print(f"   {fmt(openpyxl)}")
    print(f"   {fmt(libreoffice)}")
    print()

    expected = witan.passes and not openpyxl.passes and not libreoffice.passes
    print(f"Summary: {'expected comparison' if expected else 'unexpected result'}")
    return 0 if expected else 1


if __name__ == "__main__":
    raise SystemExit(main())
