"""Task: insert 1 new row at row 5 (shifting rows 5..10 down to 6..11). Fill the
new row 5 with Revenue=525, Cost=250. Report all formulas and cached values."""
import sys, shutil
from openpyxl import load_workbook

SRC, DST = sys.argv[1], sys.argv[2]
shutil.copy(SRC, DST)
wb = load_workbook(DST)
ws = wb["Data"]

ws.insert_rows(5, amount=1)
ws["A5"] = 525
ws["B5"] = 250
wb.save(DST)

print("=== Formulas that were supposed to shift ===")
wb2 = load_workbook(DST)
ws2 = wb2["Data"]
for addr in ("C2","C5","C6","C11", "E2","E3","E4","E5","E6", "G2","G3"):
    v = ws2[addr].value
    print(f"  {addr}  = {v!r}")

print()
print("=== Defined names ===")
for name in wb2.defined_names:
    print(f"  {name} -> {wb2.defined_names[name].attr_text!r}")

print()
print("=== Array-formula B/G refs ===")
af = ws2["G2"].value
if hasattr(af, "__dict__"):
    print(f"  G2 ArrayFormula attrs = {vars(af)}")

print()
print("=== Cached values (data_only) ===")
wb3 = load_workbook(DST, data_only=True)
ws3 = wb3["Data"]
for addr in ("C2","C5","C6","C11", "E2","E3","E4","E5","E6", "G2","G3","G11"):
    print(f"  {addr}  = {ws3[addr].value!r}")
