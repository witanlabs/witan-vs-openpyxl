"""Retest the 14 README cases using openpyxl plus a LibreOffice Calc resave.

The pairing tested here is intentionally mechanical:

1. Run the openpyxl side of the case, where that case has an authoring step.
2. If a workbook exists, open/recalculate/save it with LibreOffice headless.
3. Inspect the post-LibreOffice workbook with openpyxl and/or OOXML checks.

Cases that are pure openpyxl API failures remain failures unless LibreOffice can
reasonably be used as a pre/post-processing step for that task.
"""
from __future__ import annotations

import os
import re
import shutil
import subprocess
import sys
import tempfile
import xml.etree.ElementTree as ET
import zipfile
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.formula import Tokenizer


ROOT = Path(__file__).resolve().parents[1]
FIXTURES = ROOT / "fixtures"
SCRIPTS = ROOT / "scripts"
OUT = ROOT / "outputs" / "openpyxl_libreoffice"
SOFFICE = Path(os.environ.get("SOFFICE", "/opt/homebrew/bin/soffice"))

NS = {
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "c": "http://schemas.openxmlformats.org/drawingml/2006/chart",
}


@dataclass
class Result:
    case: int
    status: str
    task: str
    evidence: str


def run(cmd: list[str], *, check: bool = True) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        cmd,
        cwd=ROOT,
        check=check,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
    )


def py(script: str, *args: Path | str, check: bool = True) -> subprocess.CompletedProcess[str]:
    return run([sys.executable, str(SCRIPTS / script), *map(str, args)], check=check)


def resave_with_libreoffice(src: Path, stem: str) -> Path:
    if not SOFFICE.exists():
        raise RuntimeError(f"LibreOffice not found at {SOFFICE}; set SOFFICE=/path/to/soffice")

    work = OUT / "_lo_work" / stem
    converted = OUT / "_lo_out" / stem
    profile = OUT / "_lo_profile" / stem
    shutil.rmtree(work, ignore_errors=True)
    shutil.rmtree(converted, ignore_errors=True)
    shutil.rmtree(profile, ignore_errors=True)
    work.mkdir(parents=True)
    converted.mkdir(parents=True)
    profile.mkdir(parents=True)

    local = work / src.name
    shutil.copy(src, local)
    profile_uri = profile.resolve().as_uri()
    run(
        [
            str(SOFFICE),
            "--headless",
            f"-env:UserInstallation={profile_uri}",
            "--convert-to",
            "xlsx",
            "--outdir",
            str(converted),
            str(local),
        ]
    )
    out = converted / f"{local.stem}.xlsx"
    if not out.exists():
        raise RuntimeError(f"LibreOffice did not produce {out}")
    return out


def sheet_formula_values(path: Path) -> dict[str, str | None]:
    out: dict[str, str | None] = {}
    wb = load_workbook(path, data_only=True)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                out[f"{ws.title}!{cell.coordinate}"] = cell.value
    return out


def formulas(path: Path, sheet: str, addrs: list[str]) -> dict[str, object]:
    wb = load_workbook(path, data_only=False)
    return {addr: wb[sheet][addr].value for addr in addrs}


def cached(path: Path, sheet: str, addrs: list[str]) -> dict[str, object]:
    wb = load_workbook(path, data_only=True)
    return {addr: wb[sheet][addr].value for addr in addrs}


def zip_text(path: Path, member: str) -> str:
    with zipfile.ZipFile(path) as z:
        return z.read(member).decode("utf-8")


def zip_names(path: Path) -> list[str]:
    with zipfile.ZipFile(path) as z:
        return z.namelist()


def merge_ranges(path: Path) -> list[str]:
    xml = zip_text(path, "xl/worksheets/sheet1.xml")
    return re.findall(r'<mergeCell[^>]+ref="([^"]+)"', xml)


def chart_series_count(path: Path) -> int:
    chart_members = [n for n in zip_names(path) if n.startswith("xl/charts/chart")]
    if not chart_members:
        return 0
    root = ET.fromstring(zip_text(path, chart_members[0]))
    return len(root.findall(".//c:ser", NS))


def data_table_ref(path: Path) -> str | None:
    xml = zip_text(path, "xl/worksheets/sheet1.xml")
    match = re.search(r'<f[^>]*t="dataTable"[^>]*ref="([^"]+)"', xml)
    return match.group(1) if match else None


def whitespace_runs_ok(path: Path) -> bool:
    members = zip_names(path)
    xml = zip_text(path, "xl/worksheets/sheet1.xml")
    if "xl/sharedStrings.xml" in members:
        xml += zip_text(path, "xl/sharedStrings.xml")
    runs = re.findall(r'<t([^>]*)>([^<]*)</t>', xml)
    interesting = [
        (attrs, body)
        for attrs, body in runs
        if body and (body.strip() == "" or body.startswith(" ") or body.endswith(" "))
    ]
    return bool(interesting) and all('xml:space="preserve"' in attrs for attrs, _ in interesting)


def actual_borders(path: Path) -> dict[str, str]:
    sheet_xml = re.sub(r"<(/?)x:", r"<\1", zip_text(path, "xl/worksheets/sheet1.xml"))
    styles_xml = re.sub(r"<(/?)x:", r"<\1", zip_text(path, "xl/styles.xml"))
    cell_xfs = re.search(r"<cellXfs[^>]*>(.*?)</cellXfs>", styles_xml, re.DOTALL).group(1)
    xfs = re.findall(r'<xf[^/]*?borderId="(\d+)"', cell_xfs)
    borders_block = re.search(r"<borders[^>]*>(.*?)</borders>", styles_xml, re.DOTALL).group(1)
    borders = re.findall(r"<border[^>]*>.*?</border>|<border[^/]*/>", borders_block, re.DOTALL)

    def summarize(border_xml: str) -> str:
        parts: list[str] = []
        for side in ("top", "bottom", "left", "right"):
            match = re.search(rf'<{side}\s+style="(\w+)"[^/>]*(?:>.*?</{side}>|/>)', border_xml, re.DOTALL)
            if match and match.group(1) != "none":
                color = re.search(rf"<{side}[^>]*>.*?rgb=\"(\w+)\"", border_xml, re.DOTALL)
                parts.append(f"{side}:{match.group(1)}:{color.group(1)[-6:] if color else '-'}")
        return "|".join(parts) or "none"

    out: dict[str, str] = {}
    for ref in ("A1", "B1", "A2", "B2", "D1"):
        match = re.search(rf'<c r="{ref}"[^>]*s="(\d+)"', sheet_xml)
        style = int(match.group(1)) if match else None
        border_id = int(xfs[style]) if style is not None and style < len(xfs) else None
        out[ref] = summarize(borders[border_id]) if border_id is not None else "missing"
    return out


def openpyxl_borders(path: Path) -> dict[str, str]:
    wb = load_workbook(path)
    ws = wb["Data"]

    def side(s) -> str:
        if s is None or s.style is None:
            return "none"
        color = s.color.rgb[-6:] if s.color and getattr(s.color, "rgb", None) else "-"
        return f"{s.style}:{color}"

    out: dict[str, str] = {}
    for ref in ("A1", "B1", "A2", "B2", "D1"):
        b = ws[ref].border
        out[ref] = "|".join(
            f"{name}:{side_obj}"
            for name, side_obj in [
                ("top", side(b.top)),
                ("bottom", side(b.bottom)),
                ("left", side(b.left)),
                ("right", side(b.right)),
            ]
            if side_obj != "none"
        ) or "none"
    return out


def case1() -> Result:
    dst = OUT / "case1_openpyxl.xlsx"
    py("case1_openpyxl.py", FIXTURES / "pricing.xlsx", dst)
    lo = resave_with_libreoffice(dst, "case1")
    value = load_workbook(lo, data_only=True)["Summary"]["E23"].value
    ok = isinstance(value, (int, float)) and abs(value - 68407.8454990259) < 1e-6
    return Result(1, "PASS" if ok else "FAIL", "What-if NPV", f"Summary!E23 after LO = {value!r}")


def case2() -> Result:
    dst = OUT / "case2_openpyxl.xlsx"
    py("case2_openpyxl.py", FIXTURES / "circular.xlsx", dst)
    lo = resave_with_libreoffice(dst, "case2")
    vals = cached(lo, "Model", ["B3", "B4", "B7"])
    ok = isinstance(vals["B7"], (int, float)) and abs(vals["B7"] - 35000) < 1e-3
    return Result(2, "PASS" if ok else "FAIL", "Iterative circular calculation", f"Model values after LO = {vals}")


def case3() -> Result:
    lo = resave_with_libreoffice(FIXTURES / "formulas.xls", "case3")
    wb = load_workbook(lo, data_only=True)
    value = wb[wb.sheetnames[0]]["B3"].value
    return Result(3, "PASS", "Read legacy .xls", f"LibreOffice converted .xls to .xlsx; B3 = {value!r}")


def case4() -> Result:
    dst = OUT / "case4_openpyxl.xlsx"
    py("case4_openpyxl.py", FIXTURES / "review.xlsx", dst)
    lo = resave_with_libreoffice(dst, "case4")
    parts = [n for n in zip_names(lo) if "thread" in n.lower() or "person" in n.lower() or "comment" in n.lower()]
    threaded = [n for n in parts if "threaded" in n.lower() or "person" in n.lower()]
    ok = bool(threaded)
    return Result(4, "PASS" if ok else "FAIL", "Preserve/add threaded comments", f"comment-related parts after LO = {parts}")


def case5() -> Result:
    dst = OUT / "case5_openpyxl.xlsx"
    py("case5_openpyxl.py", FIXTURES / "report.xlsx", dst)
    lo = resave_with_libreoffice(dst, "case5")
    vals = cached(lo, "Summary", [f"D{r}" for r in range(2, 7)])
    expected = ["Food", "Rent", "Travel", "Supplies"]
    got = [vals[f"D{r}"] for r in range(2, 6)]
    return Result(5, "PASS" if got == expected else "FAIL", "Write dynamic-array UNIQUE/FILTER", f"Summary!D2:D6 after LO = {vals}")


def case6() -> Result:
    dst = OUT / "case6_openpyxl.xlsx"
    py("case6_openpyxl.py", dst)
    lo = resave_with_libreoffice(dst, "case6")
    count = chart_series_count(lo)
    return Result(6, "PASS" if count == 1 else "FAIL", "Single-series LineChart", f"chart series count after LO = {count}")


def case7() -> Result:
    lo = resave_with_libreoffice(FIXTURES / "report_spillref.xlsx", "case7")
    wb = load_workbook(lo, data_only=False)
    ws = wb["Summary"]
    failures = []
    for addr in ("F2", "G2", "H2"):
        try:
            Tokenizer(ws[addr].value)
        except Exception as exc:
            failures.append(f"{addr}: {type(exc).__name__}")
    cached_vals = cached(lo, "Summary", ["F2", "G2", "H2"])
    ok = not failures
    return Result(7, "PASS" if ok else "FAIL", "Parse/evaluate A1# spill references", f"Tokenizer failures = {failures}; cached after LO = {cached_vals}")


def case8() -> Result:
    dst = OUT / "case8_openpyxl_extend.xlsx"
    py("case8_openpyxl_extend.py", FIXTURES / "sensitivity2d.xlsx", dst)
    lo = resave_with_libreoffice(dst, "case8")
    vals = cached(lo, "Model", [f"{col}7" for col in "DEFGHI"])
    ref = data_table_ref(lo)
    ok = ref == "E2:I7" and all(vals[f"{col}7"] is not None for col in "EFGHI")
    return Result(8, "PASS" if ok else "FAIL", "Extend two-variable data table", f"dataTable ref = {ref!r}; row 7 = {vals}")


def case9() -> Result:
    dst = OUT / "case9_openpyxl.xlsx"
    proc = py("case9_openpyxl.py", dst, check=False)
    documented_paths_failed = "Path 1" in proc.stdout and "Path 2" in proc.stdout and "FAILED" in proc.stdout
    valid_primary = False
    if dst.exists():
        try:
            with zipfile.ZipFile(dst) as z:
                valid_primary = "[Content_Types].xml" in z.namelist() and "xl/workbook.xml" in z.namelist()
        except zipfile.BadZipFile:
            valid_primary = False
    ok = valid_primary and not documented_paths_failed
    evidence = "documented paths failed; only the space-separated workaround saved"
    return Result(9, "PASS" if ok else "FAIL", "Discontiguous conditional formatting via documented paths", f"{evidence}; script output: {proc.stdout.strip()!r}")


def case10() -> Result:
    dst = OUT / "case10_openpyxl.xlsx"
    py("case10_openpyxl.py", FIXTURES / "rename.xlsx", dst)
    lo = resave_with_libreoffice(dst, "case10")
    f = formulas(lo, "Summary", ["B1", "B2", "B4"])
    ok = all("Parameters!" in str(v) and "Inputs!" not in str(v) for v in f.values())
    return Result(10, "PASS" if ok else "FAIL", "Rename sheet referenced by formulas", f"formulas after LO = {f}")


def case11() -> Result:
    dst = OUT / "case11_openpyxl.xlsx"
    py("case11_openpyxl.py", FIXTURES / "shift.xlsx", dst)
    lo = resave_with_libreoffice(dst, "case11")
    f = formulas(lo, "Data", ["C5", "C6", "E2", "E5", "E6", "G2"])
    vals = cached(lo, "Data", ["C5", "C6", "E2", "E5", "E6", "G2", "G11"])
    wb = load_workbook(lo)
    name = wb.defined_names["RevenueRange"].attr_text
    ok = f["C5"] == "=A5-B5" and "A2:A11" in str(f["E2"]) and "$A$2:$A$11" in name
    return Result(11, "PASS" if ok else "FAIL", "Insert row and shift formulas/names/arrays", f"formulas = {f}; cached = {vals}; RevenueRange = {name!r}")


def case12() -> Result:
    dst = OUT / "case12_openpyxl.xlsx"
    py("case12_openpyxl.py", dst)
    lo = resave_with_libreoffice(dst, "case12")
    ok = whitespace_runs_ok(lo)
    return Result(12, "PASS" if ok else "FAIL", "Rich text whitespace-only runs", f"all leading/trailing/whitespace text runs have xml:space preserve after LO = {ok}")


def case13() -> Result:
    dst = OUT / "case13_openpyxl.xlsx"
    py("case13_openpyxl.py", dst)
    lo = resave_with_libreoffice(dst, "case13")
    ranges = merge_ranges(lo)
    return Result(13, "PASS" if len(ranges) == 1 else "FAIL", "Overlapping merges", f"merge ranges after LO = {ranges}")


def case14() -> Result:
    lo = resave_with_libreoffice(FIXTURES / "merge_borders.xlsx", "case14")
    actual = actual_borders(lo)
    seen = openpyxl_borders(lo)
    ok = actual == seen
    return Result(14, "PASS" if ok else "FAIL", "Read per-cell borders inside a merge", f"XML borders = {actual}; openpyxl sees = {seen}")


CASES = [
    case1,
    case2,
    case3,
    case4,
    case5,
    case6,
    case7,
    case8,
    case9,
    case10,
    case11,
    case12,
    case13,
    case14,
]


def main() -> int:
    shutil.rmtree(OUT, ignore_errors=True)
    OUT.mkdir(parents=True)
    results: list[Result] = []
    for fn in CASES:
        try:
            result = fn()
        except Exception as exc:
            case_no = int(fn.__name__.replace("case", ""))
            result = Result(case_no, "ERROR", "", f"{type(exc).__name__}: {exc}")
        results.append(result)
        print(f"{result.case:02d} {result.status:5s} {result.task}")
        print(f"   {result.evidence}")

    passed = sum(1 for r in results if r.status == "PASS")
    failed = sum(1 for r in results if r.status == "FAIL")
    errors = sum(1 for r in results if r.status == "ERROR")
    print()
    print(f"Summary: {passed} pass, {failed} fail, {errors} error")
    return 1 if errors else 0


if __name__ == "__main__":
    raise SystemExit(main())
