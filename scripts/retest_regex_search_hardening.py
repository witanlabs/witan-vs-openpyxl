"""Case 39: pathological regex search over workbook text.

The search pattern is a classic catastrophic-backtracking expression:

    ^(a+)+$

Witan's findCells path serializes regex matchers to a non-backtracking engine,
so the search remains bounded. A naive openpyxl workflow has no guarded search
API and commonly reaches for Python's `re`, which can hang on the same input.
LibreOffice resaving the workbook does not change that search behavior.
"""
from __future__ import annotations

import json
import os
import shlex
import shutil
import subprocess
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "outputs" / "regex_search_hardening"
SOFFICE = Path(os.environ.get("SOFFICE", "/opt/homebrew/bin/soffice"))
WITAN_CMD = shlex.split(os.environ.get("WITAN_CMD", "npx witan"))
TEXT = ("a" * 32000) + "!"
PATTERN = "^(a+)+$"


def run(cmd: list[str], timeout: int | None = None) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        cmd,
        cwd=ROOT,
        check=True,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        timeout=timeout,
    )


def make_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = TEXT
    wb.save(path)


def resave_with_libreoffice(src: Path) -> Path:
    if not SOFFICE.exists():
        raise RuntimeError(f"LibreOffice not found at {SOFFICE}; set SOFFICE=/path/to/soffice")

    work = OUT / "_lo_work"
    converted = OUT / "_lo_out"
    profile = OUT / "_lo_profile"
    shutil.rmtree(work, ignore_errors=True)
    shutil.rmtree(converted, ignore_errors=True)
    shutil.rmtree(profile, ignore_errors=True)
    work.mkdir(parents=True)
    converted.mkdir(parents=True)
    profile.mkdir(parents=True)

    local = work / src.name
    shutil.copy(src, local)
    run(
        [
            str(SOFFICE),
            "--headless",
            f"-env:UserInstallation={profile.resolve().as_uri()}",
            "--convert-to",
            "xlsx",
            "--outdir",
            str(converted),
            str(local),
        ],
        timeout=20,
    )
    out = converted / local.name
    if not out.exists():
        raise RuntimeError(f"LibreOffice did not produce {out}")
    return out


def witan_find(path: Path) -> dict[str, Any]:
    code = f'''
const start = Date.now();
const matches = await xlsx.findCells(
  wb,
  {{ source: {json.dumps(PATTERN)}, flags: "" }},
  {{ in: "Sheet1!A1", limit: 5 }}
);
return {{ matches, elapsedMs: Date.now() - start }};
'''
    proc = run([*WITAN_CMD, "xlsx", "exec", str(path), "--code", code], timeout=20)
    return json.loads(proc.stdout)


def python_re_search(path: Path) -> str:
    code = f"""
import re
from openpyxl import load_workbook
wb = load_workbook({str(path)!r}, read_only=True, data_only=True)
value = wb["Sheet1"]["A1"].value
print(bool(re.search({PATTERN!r}, value)))
"""
    try:
        proc = run([sys.executable, "-c", code], timeout=2)
    except subprocess.TimeoutExpired:
        return "timed out after 2s"
    return proc.stdout.strip()


def main() -> int:
    shutil.rmtree(OUT, ignore_errors=True)
    OUT.mkdir(parents=True)

    openpyxl_path = OUT / "case39_openpyxl.xlsx"
    make_workbook(openpyxl_path)

    witan = witan_find(openpyxl_path)
    openpyxl = python_re_search(openpyxl_path)
    libreoffice_path = resave_with_libreoffice(openpyxl_path)
    libreoffice = python_re_search(libreoffice_path)

    print("39 Pathological regex search")
    print(f"   witan findCells: {witan}")
    print(f"   openpyxl/Python re: {openpyxl}")
    print(f"   openpyxl + LibreOffice then Python re: {libreoffice}")
    print()

    unexpected = (
        witan.get("matches") != []
        or not isinstance(witan.get("elapsedMs"), int)
        or witan.get("elapsedMs", 10000) > 1000
        or openpyxl != "timed out after 2s"
        or libreoffice != "timed out after 2s"
    )
    print(f"Summary: {'expected comparison' if not unexpected else 'unexpected result'}")
    return 1 if unexpected else 0


if __name__ == "__main__":
    raise SystemExit(main())
