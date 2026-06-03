"""Case 41: setCells auto-reflows row height for multiline text.

Plain `setCells` in Witan reflows rows that need automatic height changes,
matching the Excel-style behavior for multiline text. It does not auto-fit
column widths; explicit `autoFitColumns` remains the operation for that.
openpyxl serializes the newline text without row-height calculation, while a
LibreOffice resave computes a different row height.
"""
from __future__ import annotations

import json
import os
import re
import shlex
import shutil
import subprocess
import zipfile
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "outputs" / "setcells_auto_row_height"
SOFFICE = Path(os.environ.get("SOFFICE", "/opt/homebrew/bin/soffice"))
WITAN_CMD = shlex.split(os.environ.get("WITAN_CMD", "npx witan"))
LONG_TEXT = "This is a deliberately long heading that does not auto-fit the column"
MULTILINE_TEXT = "Line 1\nLine 2\nLine 3\nLine 4"


def run(cmd: list[str], timeout: int | None = None) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        cmd,
        cwd=ROOT,
        check=True,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        timeout=timeout,
    )


def resave_with_libreoffice(src: Path) -> Path:
    if not SOFFICE.exists():
        raise RuntimeError(f"LibreOffice not found at {SOFFICE}; set SOFFICE=/path/to/soffice")

    work = OUT / "_lo_work"
    converted = OUT / "_lo_out"
    profile = OUT / "_lo_profile"
    shutil.rmtree(work, ignore_errors=True)
    shutil.rmtree(converted, ignore_errors=True)
    shutil.rmtree(profile, ignore_errors=True)
    work.mkdir(parents=True)
    converted.mkdir(parents=True)
    profile.mkdir(parents=True)

    local = work / src.name
    shutil.copy(src, local)
    run(
        [
            str(SOFFICE),
            "--headless",
            f"-env:UserInstallation={profile.resolve().as_uri()}",
            "--convert-to",
            "xlsx",
            "--outdir",
            str(converted),
            str(local),
        ],
        timeout=20,
    )
    out = converted / local.name
    if not out.exists():
        raise RuntimeError(f"LibreOffice did not produce {out}")
    return out


def witan_create() -> Path:
    dst = OUT / "case41_witan.xlsx"
    code = f'''
await xlsx.setCells(wb, [
  {{ address: "Sheet1!A1", value: {json.dumps(LONG_TEXT)} }},
  {{ address: "Sheet1!B1", value: "short" }},
  {{ address: "Sheet1!A3", value: {json.dumps(MULTILINE_TEXT)} }}
]);
return await xlsx.readRangeTsv(wb, "Sheet1!A1:B3");
'''
    run([*WITAN_CMD, "xlsx", "exec", str(dst), "--create", "--save", "--code", code], timeout=20)
    return dst


def openpyxl_create() -> Path:
    dst = OUT / "case41_openpyxl.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = LONG_TEXT
    ws["B1"] = "short"
    ws["A3"] = MULTILINE_TEXT
    wb.save(dst)
    return dst


def inspect_dimensions(path: Path) -> dict[str, Any]:
    wb = load_workbook(path)
    ws = wb["Sheet1"]
    with zipfile.ZipFile(path) as archive:
        xml = archive.read("xl/worksheets/sheet1.xml").decode("utf-8-sig")

    row3_match = re.search(r'<(?:\w+:)?row\b[^>]*\br="3"[^>]*>', xml)
    col_match = re.search(r"<(?:\w+:)?col\b[^>]*/>", xml)
    return {
        "columnAWidth": ws.column_dimensions["A"].width,
        "row3Height": ws.row_dimensions[3].height,
        "row3Xml": row3_match.group(0) if row3_match else None,
        "colXml": col_match.group(0) if col_match else None,
    }


def main() -> int:
    shutil.rmtree(OUT, ignore_errors=True)
    OUT.mkdir(parents=True)

    witan = inspect_dimensions(witan_create())
    openpyxl_path = openpyxl_create()
    openpyxl = inspect_dimensions(openpyxl_path)
    libreoffice = inspect_dimensions(resave_with_libreoffice(openpyxl_path))

    print("41 setCells auto-reflows multiline row height")
    print(f"   witan: {witan}")
    print(f"   openpyxl: {openpyxl}")
    print(f"   openpyxl + LibreOffice: {libreoffice}")
    print()

    unexpected = (
        witan["columnAWidth"] != 13.0
        or witan["row3Height"] != 65.0
        or openpyxl["columnAWidth"] != 13.0
        or openpyxl["row3Height"] is not None
        or libreoffice["columnAWidth"] != 13.0
        or libreoffice["row3Height"] == 65.0
    )
    print(f"Summary: {'expected comparison' if not unexpected else 'unexpected result'}")
    return 1 if unexpected else 0


if __name__ == "__main__":
    raise SystemExit(main())
