"""Modern Excel Lambda-family cases compared across Witan, openpyxl, and
openpyxl + LibreOffice.

Each case asks the tools to calculate/report a modern Excel formula. Witan is
the control because it evaluates the formula directly. openpyxl writes the
formula but has no calculation engine. The LibreOffice pairing is tested by
letting LibreOffice Calc open/recalculate/save the openpyxl-produced workbook.
"""
from __future__ import annotations

import json
import os
import shlex
import shutil
import subprocess
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "outputs" / "modern_excel_lambda"
SOFFICE = Path(os.environ.get("SOFFICE", "/opt/homebrew/bin/soffice"))
WITAN_CMD = shlex.split(os.environ.get("WITAN_CMD", "npx witan"))


@dataclass
class Case:
    number: int
    name: str
    excel_formula: str
    ooxml_formula: str
    expected: object
    expected_lo_error: str


CASES = [
    Case(
        15,
        "Inline LAMBDA",
        "=LAMBDA(x,x+1)(10)",
        "=_xlfn.LAMBDA(_xlpm.x,_xlpm.x+1)(10)",
        11,
        "#VALUE!",
    ),
    Case(
        16,
        "LET-bound LAMBDA",
        "=LET(f,LAMBDA(x,x+1),f(10))",
        "=_xlfn.LET(_xlpm.f,_xlfn.LAMBDA(_xlpm.x,_xlpm.x+1),_xlpm.f(10))",
        11,
        "#VALUE!",
    ),
    Case(
        17,
        "MAP with LAMBDA",
        "=MAP({1;2;3},LAMBDA(x,x*2))",
        "=_xlfn.MAP({1;2;3},_xlfn.LAMBDA(_xlpm.x,_xlpm.x*2))",
        [[2], [4], [6]],
        "#NAME?",
    ),
]


def run(cmd: list[str]) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        cmd,
        cwd=ROOT,
        check=True,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
    )


def build_openpyxl_workbook(case: Case) -> Path:
    path = OUT / f"case{case.number}_openpyxl.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Model"
    ws["A1"] = "Case"
    ws["B1"] = case.name
    ws["A2"] = "Formula"
    ws["B2"] = case.excel_formula
    ws["A4"] = "Result"
    ws["B4"] = case.ooxml_formula
    wb.save(path)
    return path


def resave_with_libreoffice(src: Path, stem: str) -> Path:
    if not SOFFICE.exists():
        raise RuntimeError(f"LibreOffice not found at {SOFFICE}; set SOFFICE=/path/to/soffice")

    work = OUT / "_lo_work" / stem
    converted = OUT / "_lo_out" / stem
    profile = OUT / "_lo_profile" / stem
    shutil.rmtree(work, ignore_errors=True)
    shutil.rmtree(converted, ignore_errors=True)
    shutil.rmtree(profile, ignore_errors=True)
    work.mkdir(parents=True)
    converted.mkdir(parents=True)
    profile.mkdir(parents=True)

    local = work / src.name
    shutil.copy(src, local)
    run(
        [
            str(SOFFICE),
            "--headless",
            f"-env:UserInstallation={profile.resolve().as_uri()}",
            "--convert-to",
            "xlsx",
            "--outdir",
            str(converted),
            str(local),
        ]
    )
    out = converted / local.name
    if not out.exists():
        raise RuntimeError(f"LibreOffice did not produce {out}")
    return out


def cell_value(path: Path, data_only: bool) -> object:
    return load_workbook(path, data_only=data_only)["Model"]["B4"].value


def witan_results() -> list[dict[str, object]]:
    formulas_json = json.dumps([case.excel_formula for case in CASES])
    code = (
        'await xlsx.addSheet(wb, "Model"); '
        f"return await xlsx.evaluateFormulas(wb, \"Model\", {formulas_json});"
    )
    proc = run([*WITAN_CMD, "xlsx", "exec", str(OUT / "_witan_probe.xlsx"), "--create", "--code", code])
    return json.loads(proc.stdout)


def same_value(actual: object, expected: object) -> bool:
    if isinstance(expected, list):
        return actual == expected
    if isinstance(expected, (int, float)) and isinstance(actual, (int, float)):
        return abs(actual - expected) < 1e-9
    return actual == expected


def main() -> int:
    shutil.rmtree(OUT, ignore_errors=True)
    OUT.mkdir(parents=True)

    witan = witan_results()
    unexpected = 0

    for case, witan_result in zip(CASES, witan):
        openpyxl_path = build_openpyxl_workbook(case)
        openpyxl_cached = cell_value(openpyxl_path, data_only=True)

        lo_path = resave_with_libreoffice(openpyxl_path, f"case{case.number}")
        lo_cached = cell_value(lo_path, data_only=True)
        lo_formula = cell_value(lo_path, data_only=False)

        witan_value = witan_result.get("value")
        witan_ok = not witan_result.get("error") and same_value(witan_value, case.expected)
        openpyxl_ok = same_value(openpyxl_cached, case.expected)
        lo_ok = same_value(lo_cached, case.expected)

        if not witan_ok or openpyxl_ok or lo_ok or lo_cached != case.expected_lo_error:
            unexpected += 1

        print(f"{case.number:02d} {case.name}")
        print(f"   expected: {case.expected!r}")
        print(f"   witan: {witan_value!r} error={witan_result.get('error')!r}")
        print(f"   openpyxl data_only after write: {openpyxl_cached!r}")
        print(f"   openpyxl + LibreOffice data_only: {lo_cached!r}")
        print(f"   formula after LibreOffice: {lo_formula!r}")

    print()
    print(f"Summary: {len(CASES) - unexpected} expected comparisons, {unexpected} unexpected")
    return 1 if unexpected else 0


if __name__ == "__main__":
    raise SystemExit(main())
