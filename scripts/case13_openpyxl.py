"""Task: on the Data sheet, merge A1:B2 and merge A2:C3 (overlapping)."""
import sys
from openpyxl import Workbook, load_workbook

OUT = sys.argv[1]
wb = Workbook()
ws = wb.active
ws.title = "Data"

# Seed some labels so something shows up in each cell
for r in range(1, 4):
    for c in range(1, 4):
        ws.cell(row=r, column=c, value=f"{chr(64+c)}{r}")

print("=== Apply overlapping merges ===")
try:
    ws.merge_cells("A1:B2")
    print("  merge A1:B2 → no error")
except Exception as e:
    print(f"  merge A1:B2 FAILED ({type(e).__name__}): {e}")

try:
    ws.merge_cells("A2:C3")
    print("  merge A2:C3 → no error")
except Exception as e:
    print(f"  merge A2:C3 FAILED ({type(e).__name__}): {e}")

try:
    wb.save(OUT)
    print(f"  save → OK  ({OUT})")
except Exception as e:
    print(f"  save FAILED ({type(e).__name__}): {e}")

# Inspect merged ranges in the saved file
print()
print("=== Raw <mergeCells> in saved XML ===")
import zipfile
with zipfile.ZipFile(OUT) as z:
    xml = z.read("xl/worksheets/sheet1.xml").decode()
import re
for m in re.finditer(r'<(?:mergeCell|mergeCells)[^/>]*>', xml):
    print(f"  {m.group()}")

print()
print("=== After reload ===")
wb2 = load_workbook(OUT)
print(f"  ws.merged_cells.ranges = {list(map(str, wb2['Data'].merged_cells.ranges))}")
