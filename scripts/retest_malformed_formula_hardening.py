"""Case 40: malformed formula authoring guard.

Witan validates formulas during setCells and rejects malformed formula text
before mutating the workbook. openpyxl serializes the malformed formula string
as-is; LibreOffice then rewrites it into a different formula and caches a
#VALUE! result rather than surfacing a precise authoring error.
"""
from __future__ import annotations

import json
import os
import shlex
import shutil
import subprocess
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "outputs" / "malformed_formula_hardening"
SOFFICE = Path(os.environ.get("SOFFICE", "/opt/homebrew/bin/soffice"))
WITAN_CMD = shlex.split(os.environ.get("WITAN_CMD", "npx witan"))
FORMULA = "=LAMBDA(x,x+1)(2"


def run(cmd: list[str], timeout: int | None = None) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        cmd,
        cwd=ROOT,
        check=True,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        timeout=timeout,
    )


def resave_with_libreoffice(src: Path) -> Path:
    if not SOFFICE.exists():
        raise RuntimeError(f"LibreOffice not found at {SOFFICE}; set SOFFICE=/path/to/soffice")

    work = OUT / "_lo_work"
    converted = OUT / "_lo_out"
    profile = OUT / "_lo_profile"
    shutil.rmtree(work, ignore_errors=True)
    shutil.rmtree(converted, ignore_errors=True)
    shutil.rmtree(profile, ignore_errors=True)
    work.mkdir(parents=True)
    converted.mkdir(parents=True)
    profile.mkdir(parents=True)

    local = work / src.name
    shutil.copy(src, local)
    run(
        [
            str(SOFFICE),
            "--headless",
            f"-env:UserInstallation={profile.resolve().as_uri()}",
            "--convert-to",
            "xlsx",
            "--outdir",
            str(converted),
            str(local),
        ],
        timeout=20,
    )
    out = converted / local.name
    if not out.exists():
        raise RuntimeError(f"LibreOffice did not produce {out}")
    return out


def witan_attempt() -> dict[str, Any]:
    dst = OUT / "case40_witan.xlsx"
    code = f'''
await xlsx.setCells(wb, [{{ address: "Sheet1!A1", value: 1 }}]);
try {{
  await xlsx.setCells(wb, [{{ address: "Sheet1!A1", formula: {json.dumps(FORMULA)} }}]);
}} catch (error) {{
  const cell = await xlsx.readCell(wb, "Sheet1!A1");
  return {{ caught: String(error && error.message || error), cell }};
}}
const cell = await xlsx.readCell(wb, "Sheet1!A1");
return {{ caught: null, cell }};
'''
    proc = run([*WITAN_CMD, "xlsx", "exec", str(dst), "--create", "--save", "--code", code], timeout=20)
    return json.loads(proc.stdout)


def openpyxl_create() -> Path:
    dst = OUT / "case40_openpyxl.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = FORMULA
    wb.save(dst)
    return dst


def read_a1(path: Path) -> dict[str, Any]:
    formula_wb = load_workbook(path, data_only=False)
    value_wb = load_workbook(path, data_only=True)
    return {
        "formula": formula_wb["Sheet1"]["A1"].value,
        "value": value_wb["Sheet1"]["A1"].value,
    }


def main() -> int:
    shutil.rmtree(OUT, ignore_errors=True)
    OUT.mkdir(parents=True)

    witan = witan_attempt()
    openpyxl_path = openpyxl_create()
    openpyxl = read_a1(openpyxl_path)
    libreoffice_path = resave_with_libreoffice(openpyxl_path)
    libreoffice = read_a1(libreoffice_path)

    print("40 Malformed formula authoring guard")
    print(f"   witan: {witan}")
    print(f"   openpyxl: {openpyxl}")
    print(f"   openpyxl + LibreOffice: {libreoffice}")
    print()

    unexpected = (
        not witan.get("caught")
        or witan.get("cell", {}).get("value") != 1
        or openpyxl.get("formula") != FORMULA
        or openpyxl.get("value") is not None
        or libreoffice.get("formula") == FORMULA
        or libreoffice.get("value") != "#VALUE!"
    )
    print(f"Summary: {'expected comparison' if not unexpected else 'unexpected result'}")
    return 1 if unexpected else 0


if __name__ == "__main__":
    raise SystemExit(main())
