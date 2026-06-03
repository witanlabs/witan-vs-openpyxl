"""Case 38: pathological regex formula evaluation.

The formula uses a classic catastrophic-backtracking pattern:

    ^(a+)+$

against a long non-matching string. Witan evaluates REGEXTEST through a
non-backtracking regex path, so the calculation remains bounded. openpyxl
cannot calculate the formula, and LibreOffice 26.2.1.2 does not support
REGEXTEST in this probe.
"""
from __future__ import annotations

import json
import os
import shlex
import shutil
import subprocess
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "outputs" / "pathological_regex_formula"
SOFFICE = Path(os.environ.get("SOFFICE", "/opt/homebrew/bin/soffice"))
WITAN_CMD = shlex.split(os.environ.get("WITAN_CMD", "npx witan"))
TEXT = ("a" * 2000) + "!"
PATTERN = "^(a+)+$"


def run(cmd: list[str], timeout: int | None = None) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        cmd,
        cwd=ROOT,
        check=True,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        timeout=timeout,
    )


def resave_with_libreoffice(src: Path) -> Path | None:
    if not SOFFICE.exists():
        raise RuntimeError(f"LibreOffice not found at {SOFFICE}; set SOFFICE=/path/to/soffice")

    work = OUT / "_lo_work"
    converted = OUT / "_lo_out"
    profile = OUT / "_lo_profile"
    shutil.rmtree(work, ignore_errors=True)
    shutil.rmtree(converted, ignore_errors=True)
    shutil.rmtree(profile, ignore_errors=True)
    work.mkdir(parents=True)
    converted.mkdir(parents=True)
    profile.mkdir(parents=True)

    local = work / src.name
    shutil.copy(src, local)
    try:
        run(
            [
                str(SOFFICE),
                "--headless",
                f"-env:UserInstallation={profile.resolve().as_uri()}",
                "--convert-to",
                "xlsx",
                "--outdir",
                str(converted),
                str(local),
            ],
            timeout=20,
        )
    except subprocess.TimeoutExpired:
        return None
    out = converted / local.name
    if not out.exists():
        raise RuntimeError(f"LibreOffice did not produce {out}")
    return out


def witan_create() -> dict[str, Any]:
    dst = OUT / "case38_witan.xlsx"
    code = f'''
await xlsx.setCells(wb, [
  {{address:"Sheet1!A1", value:{json.dumps(TEXT)}}},
  {{address:"Sheet1!B1", value:{json.dumps(PATTERN)}}},
  {{address:"Sheet1!C1", formula:"=REGEXTEST(A1,B1)"}}
]);
const start = Date.now();
const cell = await xlsx.readCell(wb, "Sheet1!C1");
return {{formula: cell.formula, value: cell.value, elapsedMs: Date.now() - start}};
'''
    proc = run([*WITAN_CMD, "xlsx", "exec", str(dst), "--create", "--save", "--code", code], timeout=20)
    return json.loads(proc.stdout)


def openpyxl_create() -> Path:
    dst = OUT / "case38_openpyxl.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = TEXT
    ws["B1"] = PATTERN
    ws["C1"] = "=REGEXTEST(A1,B1)"
    wb.save(dst)
    return dst


def read_cell(path: Path) -> dict[str, Any]:
    formula_wb = load_workbook(path, data_only=False)
    value_wb = load_workbook(path, data_only=True)
    return {
        "formula": formula_wb["Sheet1"]["C1"].value,
        "value": value_wb["Sheet1"]["C1"].value,
    }


def main() -> int:
    shutil.rmtree(OUT, ignore_errors=True)
    OUT.mkdir(parents=True)

    witan = witan_create()
    openpyxl_path = openpyxl_create()
    openpyxl = read_cell(openpyxl_path)
    libreoffice_path = resave_with_libreoffice(openpyxl_path)
    libreoffice = "timed out after 20s" if libreoffice_path is None else read_cell(libreoffice_path)

    print("38 Pathological regex formula")
    print(f"   witan: {witan}")
    print(f"   openpyxl: {openpyxl}")
    print(f"   openpyxl + LibreOffice: {libreoffice}")
    print()

    unexpected = (
        witan.get("value") is not False
        or not isinstance(witan.get("elapsedMs"), int)
        or witan.get("elapsedMs", 10000) > 1000
        or openpyxl.get("value") is not None
        or (isinstance(libreoffice, dict) and libreoffice.get("value") is False)
    )
    print(f"Summary: {'expected comparison' if not unexpected else 'unexpected result'}")
    return 1 if unexpected else 0


if __name__ == "__main__":
    raise SystemExit(main())
