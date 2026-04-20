"""Task: in Summary!D2 enter =UNIQUE(FILTER(Raw!A2:A13, Raw!B2:B13>0)) as a dynamic array,
then report the spill values from Summary!D2:D10."""
import sys, shutil
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator

SRC, DST = sys.argv[1], sys.argv[2]
shutil.copy(SRC, DST)

wb = load_workbook(DST)
ws = wb["Summary"]

# Naive agent attempt: assign as a string formula.
ws["D2"] = "=UNIQUE(FILTER(Raw!A2:A13, Raw!B2:B13>0))"
wb.save(DST)

# Attempt to read back
wb2 = load_workbook(DST, data_only=True)
ws2 = wb2["Summary"]
print("openpyxl Summary!D2:D10 (data_only) =")
for r in range(2, 11):
    print(f"  D{r} = {ws2.cell(row=r, column=4).value!r}")

wb3 = load_workbook(DST)
print("Summary!D2 formula =", wb3["Summary"]["D2"].value)
