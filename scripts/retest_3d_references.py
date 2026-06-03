"""Case 19: recalculate a formula with a 3D sheet reference.

The workbook contains Jan, Feb, Mar, and Summary sheets. Summary!B1 uses
=SUM(Jan:Mar!B2). The task changes Feb!B2 from 200 to 250 and reports the
updated quarterly total.
"""
from __future__ import annotations

import json
import os
import shlex
import shutil
import subprocess
from pathlib import Path

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "outputs" / "3d_references"
SOFFICE = Path(os.environ.get("SOFFICE", "/opt/homebrew/bin/soffice"))
WITAN_CMD = shlex.split(os.environ.get("WITAN_CMD", "npx witan"))
EXPECTED = 650


def run(cmd: list[str]) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        cmd,
        cwd=ROOT,
        check=True,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
    )


def build_fixture(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Jan"
    ws["A1"] = "Revenue"
    ws["B2"] = 100
    for name, value in [("Feb", 200), ("Mar", 300)]:
        sheet = wb.create_sheet(name)
        sheet["A1"] = "Revenue"
        sheet["B2"] = value
    summary = wb.create_sheet("Summary")
    summary["A1"] = "Quarter revenue"
    summary["B1"] = "=SUM(Jan:Mar!B2)"
    wb.save(path)


def resave_with_libreoffice(src: Path) -> Path:
    if not SOFFICE.exists():
        raise RuntimeError(f"LibreOffice not found at {SOFFICE}; set SOFFICE=/path/to/soffice")

    work = OUT / "_lo_work"
    converted = OUT / "_lo_out"
    profile = OUT / "_lo_profile"
    shutil.rmtree(work, ignore_errors=True)
    shutil.rmtree(converted, ignore_errors=True)
    shutil.rmtree(profile, ignore_errors=True)
    work.mkdir(parents=True)
    converted.mkdir(parents=True)
    profile.mkdir(parents=True)

    local = work / src.name
    shutil.copy(src, local)
    run(
        [
            str(SOFFICE),
            "--headless",
            f"-env:UserInstallation={profile.resolve().as_uri()}",
            "--convert-to",
            "xlsx",
            "--outdir",
            str(converted),
            str(local),
        ]
    )
    out = converted / local.name
    if not out.exists():
        raise RuntimeError(f"LibreOffice did not produce {out}")
    return out


def run_witan(fixture: Path) -> object:
    dst = OUT / "case19_witan.xlsx"
    shutil.copy(fixture, dst)
    code = (
        'const r = await xlsx.setCells(wb, [{address:"Feb!B2", value:250}]); '
        'return {touched:r.touched, summary: await xlsx.readCell(wb, "Summary!B1")};'
    )
    proc = run([*WITAN_CMD, "xlsx", "exec", str(dst), "--code", code, "--save"])
    return json.loads(proc.stdout)["summary"]["value"]


def run_openpyxl(fixture: Path) -> Path:
    dst = OUT / "case19_openpyxl.xlsx"
    shutil.copy(fixture, dst)
    wb = load_workbook(dst)
    wb["Feb"]["B2"] = 250
    wb.save(dst)
    return dst


def summary_value(path: Path, data_only: bool) -> object:
    return load_workbook(path, data_only=data_only)["Summary"]["B1"].value


def main() -> int:
    shutil.rmtree(OUT, ignore_errors=True)
    OUT.mkdir(parents=True)
    fixture = OUT / "case19_fixture.xlsx"
    build_fixture(fixture)

    witan = run_witan(fixture)
    openpyxl_path = run_openpyxl(fixture)
    openpyxl_cached = summary_value(openpyxl_path, data_only=True)
    lo_path = resave_with_libreoffice(openpyxl_path)
    lo_cached = summary_value(lo_path, data_only=True)
    lo_formula = summary_value(lo_path, data_only=False)

    print("19 3D reference recalculation")
    print(f"   expected: {EXPECTED}")
    print(f"   witan: {witan!r}")
    print(f"   openpyxl data_only after write: {openpyxl_cached!r}")
    print(f"   openpyxl + LibreOffice data_only: {lo_cached!r}")
    print(f"   formula after LibreOffice: {lo_formula!r}")
    print()

    unexpected = witan != EXPECTED or openpyxl_cached == EXPECTED or lo_cached != EXPECTED
    print(f"Summary: {'expected comparison' if not unexpected else 'unexpected result'}")
    return 1 if unexpected else 0


if __name__ == "__main__":
    raise SystemExit(main())
