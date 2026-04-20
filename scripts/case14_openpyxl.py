"""Task: read each cell's border from the witan-built fixture and report
what openpyxl sees vs what the XML actually stores."""
import sys, zipfile, re
from openpyxl import load_workbook

SRC = sys.argv[1]

# 1. What the XML says for each cell's style index and the matching border
with zipfile.ZipFile(SRC) as z:
    sheet_xml = z.read("xl/worksheets/sheet1.xml").decode()
    styles_xml = z.read("xl/styles.xml").decode()

# Strip "x:" namespace prefix used by witan
styles_xml = re.sub(r'<(/?)x:', r'<\1', styles_xml)
sheet_xml  = re.sub(r'<(/?)x:', r'<\1', sheet_xml)

# cellXfs gives us cell-style-index -> border-index
cellXfs_block = re.search(r'<cellXfs[^>]*>(.*?)</cellXfs>', styles_xml, re.DOTALL).group(1)
xfs = re.findall(r'<xf[^/]*?borderId="(\d+)"', cellXfs_block)

# borders block
borders_block = re.search(r'<borders[^>]*>(.*?)</borders>', styles_xml, re.DOTALL).group(1)
borders = re.findall(r'<border[^>]*>.*?</border>|<border[^/]*/>', borders_block, re.DOTALL)

def summarize_border(b):
    parts = []
    for side in ("top","bottom","left","right"):
        m = re.search(rf'<{side}\s+style="(\w+)"[^/>]*(?:>.*?</{side}>|/>)', b, re.DOTALL)
        if m and m.group(1) != "none":
            style = m.group(1)
            col = re.search(rf'<{side}[^>]*>.*?rgb="(\w+)"', b, re.DOTALL)
            col = col.group(1)[-6:] if col else "-"
            parts.append(f"{side}={style}({col})")
    return ", ".join(parts) or "(all none)"

print("=== XML view ===")
for ref in ("A1","B1","A2","B2","D1"):
    m = re.search(rf'<c r="{ref}"[^>]*s="(\d+)"', sheet_xml)
    if not m:
        m = re.search(rf'<(?:x:)?c r="{ref}"[^>]*s="(\d+)"', sheet_xml)
    s = int(m.group(1)) if m else None
    border_idx = int(xfs[s]) if s is not None and s < len(xfs) else None
    b_summary = summarize_border(borders[border_idx]) if border_idx is not None and border_idx < len(borders) else "?"
    print(f"  {ref}  cellXf={s}  borderId={border_idx}  [{b_summary}]")

print()
print("=== openpyxl .border ===")
wb = load_workbook(SRC)
ws = wb["Data"]
for ref in ("A1","B1","A2","B2","D1"):
    c = ws[ref]
    b = c.border
    def side_str(s):
        if s is None or s.style is None:
            return "none"
        col = s.color.rgb if s.color and hasattr(s.color, "rgb") else "-"
        return f"{s.style}({col and col[-6:]})"
    print(f"  {ref}  type={type(c).__name__:10s}  top={side_str(b.top)}  bottom={side_str(b.bottom)}  left={side_str(b.left)}  right={side_str(b.right)}")
