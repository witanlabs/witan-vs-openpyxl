"""Quick sanity checks for the other three structural edits: delete row, insert col, delete col."""
import sys, shutil
from openpyxl import load_workbook

SRC = sys.argv[1]

def fresh(op_name):
    path = f"outputs/case12_op_{op_name}.xlsx"
    shutil.copy(SRC, path)
    return load_workbook(path), path

for op_name, op in [
    ("delete_row",   lambda ws: ws.delete_rows(5, 1)),
    ("insert_col",   lambda ws: ws.insert_cols(2, 1)),  # insert between A and B
    ("delete_col",   lambda ws: ws.delete_cols(2, 1)),
]:
    print(f"=== {op_name} ===")
    wb, path = fresh(op_name)
    ws = wb["Data"]
    op(ws)
    wb.save(path)
    wb2 = load_workbook(path)
    ws2 = wb2["Data"]
    # Print a handful of key cells
    for addr in ("C2","C5","C10", "E2","E3","E5","E6", "G2"):
        v = ws2[addr].value
        print(f"  {addr} = {v!r}")
    print(f"  RevenueRange -> {wb2.defined_names['RevenueRange'].attr_text!r}")
    print()
